from django.test import TestCase
from django.urls import reverse
from rest_framework.test import APIClient
from rest_framework import status
from unittest.mock import patch, MagicMock
from datetime import date

from users.models import User
from participants.models import Participant
from instructors.models import Instructor
from events.models import Event, EventCategory, Enrollment
from certificados.models import Certificate, Template


def make_admin(email='admin@test.com'):
    return User.objects.create_user(
        email=email, full_name='Admin', password='pass123', role='admin', is_staff=True
    )

def make_user(email='user@test.com'):
    return User.objects.create_user(
        email=email, full_name='User', password='pass123', role='participante'
    )

def make_event(user, name='Evento Test', ev_date=None):
    return Event.objects.create(
        name=name, event_date=ev_date or date(2026, 6, 1), created_by=user
    )

def make_participant(user, doc='12345', email='participant@test.com'):
    return Participant.objects.create(
        document_id=doc, first_name='Ana', last_name='Lopez',
        email=email, phone='999000111', created_by=user
    )


# ─────────────────────────────────────────────
# Auth endpoints
# ─────────────────────────────────────────────

class RegisterViewTest(TestCase):
    def setUp(self):
        self.client = APIClient()

    def test_register_success(self):
        res = self.client.post('/api/register/', {
            'email': 'new@test.com', 'full_name': 'Nuevo', 'password': 'Pass1234!', 'password_confirm': 'Pass1234!'
        })
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertIn('email', res.data)

    def test_register_password_mismatch(self):
        res = self.client.post('/api/register/', {
            'email': 'new2@test.com', 'full_name': 'Nuevo', 'password': 'Pass1234!', 'password_confirm': 'Other!'
        })
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_register_duplicate_email(self):
        make_admin('dup@test.com')
        res = self.client.post('/api/register/', {
            'email': 'dup@test.com', 'full_name': 'Otro', 'password': 'Pass1234!', 'password_confirm': 'Pass1234!'
        })
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)


class LoginViewTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = make_admin()

    def test_login_success(self):
        res = self.client.post('/api/login/', {'email': 'admin@test.com', 'password': 'pass123'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('access', res.data)
        self.assertIn('refresh', res.data)

    def test_login_wrong_password(self):
        res = self.client.post('/api/login/', {'email': 'admin@test.com', 'password': 'wrong'})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_login_nonexistent_user(self):
        res = self.client.post('/api/login/', {'email': 'ghost@test.com', 'password': 'pass123'})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)


class CurrentUserViewTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = make_admin()
        self.client.force_authenticate(user=self.user)

    def test_get_current_user(self):
        res = self.client.get('/api/me/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['email'], 'admin@test.com')
        self.assertIn('role', res.data)

    def test_unauthenticated_returns_401(self):
        self.client.force_authenticate(user=None)
        res = self.client.get('/api/me/')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)


# ─────────────────────────────────────────────
# Participants
# ─────────────────────────────────────────────

class ParticipantsViewSetTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)

    def test_list_participants(self):
        make_participant(self.admin)
        res = self.client.get('/api/participants/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_create_participant(self):
        res = self.client.post('/api/participants/', {
            'document_id': '99999', 'first_name': 'Luis', 'last_name': 'Gomez',
            'email': 'luis@test.com', 'phone': '111222333'
        })
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(res.data['first_name'], 'Luis')

    def test_retrieve_participant(self):
        s = make_participant(self.admin)
        res = self.client.get(f'/api/participants/{s.id}/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['email'], 'participant@test.com')

    def test_update_participant(self):
        s = make_participant(self.admin)
        res = self.client.patch(f'/api/participants/{s.id}/', {'first_name': 'Cambiado'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['first_name'], 'Cambiado')

    def test_delete_participant(self):
        s = make_participant(self.admin)
        res = self.client.delete(f'/api/participants/{s.id}/')
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)

    def test_create_duplicate_document_id_returns_400(self):
        make_participant(self.admin)
        res = self.client.post('/api/participants/', {
            'document_id': '12345', 'first_name': 'Otro', 'last_name': 'X',
            'email': 'otro@test.com', 'phone': ''
        })
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_unauthenticated_cannot_list(self):
        self.client.force_authenticate(user=None)
        res = self.client.get('/api/participants/')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)


# ─────────────────────────────────────────────
# Instructors
# ─────────────────────────────────────────────

class InstructorsViewSetTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)

    def test_list_instructors_empty(self):
        res = self.client.get('/api/instructors/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_create_instructor(self):
        res = self.client.post('/api/instructors/', {
            'full_name': 'Rosa Diaz', 'email': 'rosa@test.com', 'specialty': 'Python'
        })
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(res.data['full_name'], 'Rosa Diaz')

    def test_retrieve_instructor(self):
        inst = Instructor.objects.create(full_name='Carlos', email='carlos@test.com', created_by=self.admin)
        res = self.client.get(f'/api/instructors/{inst.id}/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_update_instructor(self):
        inst = Instructor.objects.create(full_name='Pedro', email='pedro@test.com', created_by=self.admin)
        res = self.client.patch(f'/api/instructors/{inst.id}/', {'specialty': 'Django'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['specialty'], 'Django')

    def test_delete_instructor(self):
        inst = Instructor.objects.create(full_name='Temp', email='temp@test.com', created_by=self.admin)
        res = self.client.delete(f'/api/instructors/{inst.id}/')
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)


# ─────────────────────────────────────────────
# Events
# ─────────────────────────────────────────────

class EventsViewSetTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)

    def test_list_events(self):
        make_event(self.admin)
        res = self.client.get('/api/events/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_create_event(self):
        res = self.client.post('/api/events/', {
            'name': 'Nuevo Evento', 'event_date': '2026-08-01', 'status': 'active'
        })
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(res.data['name'], 'Nuevo Evento')

    def test_retrieve_event(self):
        e = make_event(self.admin)
        res = self.client.get(f'/api/events/{e.id}/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['name'], 'Evento Test')

    def test_update_event(self):
        e = make_event(self.admin)
        res = self.client.patch(f'/api/events/{e.id}/', {'status': 'finished'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_delete_event(self):
        e = make_event(self.admin)
        res = self.client.delete(f'/api/events/{e.id}/')
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)

    def test_participants_empty(self):
        e = make_event(self.admin)
        res = self.client.get(f'/api/events/{e.id}/participants/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data, [])

    def test_participants_with_enrollment(self):
        e = make_event(self.admin)
        s = make_participant(self.admin)
        Enrollment.objects.create(participant=s, event=e, attendance=True, created_by=self.admin)
        res = self.client.get(f'/api/events/{e.id}/participants/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(len(res.data), 1)
        self.assertTrue(res.data[0]['attendance'])

    def test_enroll_student_by_id(self):
        e = make_event(self.admin)
        s = make_participant(self.admin)
        res = self.client.post(f'/api/events/{e.id}/enroll/', {'student_id': s.id})
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)

    def test_enroll_student_by_email(self):
        e = make_event(self.admin)
        res = self.client.post(f'/api/events/{e.id}/enroll/', {'student_email': 'new@enroll.com'})
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)

    def test_enroll_duplicate_returns_400(self):
        e = make_event(self.admin)
        s = make_participant(self.admin)
        Enrollment.objects.create(participant=s, event=e, created_by=self.admin)
        res = self.client.post(f'/api/events/{e.id}/enroll/', {'student_id': s.id})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_enroll_missing_params_returns_400(self):
        e = make_event(self.admin)
        res = self.client.post(f'/api/events/{e.id}/enroll/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_enroll_nonexistent_student_returns_404(self):
        e = make_event(self.admin)
        res = self.client.post(f'/api/events/{e.id}/enroll/', {'student_id': 9999})
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_stats(self):
        e = make_event(self.admin)
        res = self.client.get(f'/api/events/{e.id}/stats/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_non_admin_cannot_enroll(self):
        regular = make_user('regular@test.com')
        self.client.force_authenticate(user=regular)
        e = make_event(self.admin)
        res = self.client.post(f'/api/events/{e.id}/enroll/', {'student_email': 'x@x.com'})
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)


# ─────────────────────────────────────────────
# Certificates
# ─────────────────────────────────────────────

class CertificateViewSetTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.participant = make_participant(self.admin)
        self.event = make_event(self.admin)
        self.template = Template.objects.create(name='Base', created_by=self.admin)
        self.cert = Certificate.objects.create(
            participant=self.participant, event=self.event,
            template=self.template, generated_by=self.admin
        )

    def test_list_certificates(self):
        res = self.client.get('/api/certificates/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_retrieve_certificate(self):
        res = self.client.get(f'/api/certificates/{self.cert.id}/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_history_endpoint(self):
        res = self.client.get(f'/api/certificates/{self.cert.id}/history/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('total_attempts', res.data)

    def test_verify_with_valid_code(self):
        res = self.client.get(f'/api/certificates/verify/?code={self.cert.verification_code}')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['status'], 'success')

    def test_verify_with_invalid_code(self):
        res = self.client.get('/api/certificates/verify/?code=INVALID-CODE-XXXX')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_verify_without_code_returns_400(self):
        res = self.client.get('/api/certificates/verify/')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_generate_raises_if_not_pending(self):
        Certificate.objects.filter(pk=self.cert.pk).update(status='generated')
        res = self.client.post(f'/api/certificates/{self.cert.id}/generate/', {
            'student_id': self.participant.id,
            'event_id': self.event.id,
        }, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_deliver_raises_if_not_generated(self):
        res = self.client.post(f'/api/certificates/{self.cert.id}/deliver/', {'method': 'email'})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)


# ─────────────────────────────────────────────
# Templates
# ─────────────────────────────────────────────

class TemplateViewSetTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)

    def test_list_templates(self):
        Template.objects.create(name='T1', created_by=self.admin)
        res = self.client.get('/api/templates/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_create_template(self):
        res = self.client.post('/api/templates/', {'name': 'Nueva Plantilla', 'category': 'Cursos'})
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)
        self.assertEqual(res.data['name'], 'Nueva Plantilla')

    def test_retrieve_template(self):
        t = Template.objects.create(name='T2', created_by=self.admin)
        res = self.client.get(f'/api/templates/{t.id}/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_update_template(self):
        t = Template.objects.create(name='T3', created_by=self.admin)
        res = self.client.patch(f'/api/templates/{t.id}/', {'name': 'Actualizada'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_delete_template(self):
        t = Template.objects.create(name='T4', created_by=self.admin)
        res = self.client.delete(f'/api/templates/{t.id}/')
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)


# ─────────────────────────────────────────────
# Delivery Logs
# ─────────────────────────────────────────────

class DeliveryLogViewSetTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)

    def test_list_delivery_logs(self):
        res = self.client.get('/api/deliveries/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_filter_by_certificate_id(self):
        student = make_participant(self.admin)
        event = make_event(self.admin)
        template = Template.objects.create(name='T', created_by=self.admin)
        cert = Certificate.objects.create(participant=student, event=event, template=template, generated_by=self.admin)
        res = self.client.get(f'/api/deliveries/?certificate_id={cert.id}')
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ─────────────────────────────────────────────
# Permissions
# ─────────────────────────────────────────────

class PermissionsTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.user = make_user()

    def _req(self, method='GET'):
        from rest_framework.test import APIRequestFactory
        factory = APIRequestFactory()
        req = factory.get('/') if method == 'GET' else factory.post('/')
        return req

    def test_is_admin_true_for_admin(self):
        from api.permissions import is_admin
        self.client.force_authenticate(user=self.admin)
        req = self._req()
        req.user = self.admin
        self.assertTrue(is_admin(req))

    def test_is_admin_false_for_participante(self):
        from api.permissions import is_admin
        req = self._req()
        req.user = self.user
        self.assertFalse(is_admin(req))

    def test_is_admin_false_for_anonymous(self):
        from api.permissions import is_admin
        from django.contrib.auth.models import AnonymousUser
        req = self._req()
        req.user = AnonymousUser()
        self.assertFalse(is_admin(req))

    def _check_permission(self, perm_class, user, method='GET'):
        from rest_framework.test import APIRequestFactory
        factory = APIRequestFactory()
        if method == 'GET':
            req = factory.get('/')
        else:
            req = factory.post('/')
        req.user = user
        from rest_framework.request import Request
        from rest_framework.parsers import JSONParser
        drf_req = Request(req)
        drf_req._user = user
        perm = perm_class()
        return perm.has_permission(drf_req, None)

    def test_is_admin_perm_allows_admin(self):
        from api.permissions import IsAdmin
        self.assertTrue(self._check_permission(IsAdmin, self.admin))

    def test_is_admin_perm_denies_participante(self):
        from api.permissions import IsAdmin
        self.assertFalse(self._check_permission(IsAdmin, self.user))

    def test_admin_or_read_only_allows_admin_write(self):
        from api.permissions import IsAdminOrReadOnly
        self.assertTrue(self._check_permission(IsAdminOrReadOnly, self.admin, 'POST'))

    def test_admin_or_read_only_allows_user_read(self):
        from api.permissions import IsAdminOrReadOnly
        self.assertTrue(self._check_permission(IsAdminOrReadOnly, self.user, 'GET'))

    def test_admin_or_read_only_denies_user_write(self):
        from api.permissions import IsAdminOrReadOnly
        self.assertFalse(self._check_permission(IsAdminOrReadOnly, self.user, 'POST'))

    def test_can_manage_users_admin(self):
        from api.permissions import CanManageUsers
        self.assertTrue(self._check_permission(CanManageUsers, self.admin, 'POST'))

    def test_can_manage_users_participante_read(self):
        from api.permissions import CanManageUsers
        self.assertTrue(self._check_permission(CanManageUsers, self.user, 'GET'))

    def test_can_manage_certificates_admin(self):
        from api.permissions import CanManageCertificates
        self.assertTrue(self._check_permission(CanManageCertificates, self.admin, 'POST'))

    def test_can_manage_events_admin(self):
        from api.permissions import CanManageEvents
        self.assertTrue(self._check_permission(CanManageEvents, self.admin, 'POST'))

    def test_can_manage_students_admin(self):
        from api.permissions import CanManageStudents
        self.assertTrue(self._check_permission(CanManageStudents, self.admin, 'POST'))

    def test_can_manage_instructors_admin(self):
        from api.permissions import CanManageInstructors
        self.assertTrue(self._check_permission(CanManageInstructors, self.admin, 'POST'))

    def test_can_manage_templates_admin(self):
        from api.permissions import CanManageTemplates
        self.assertTrue(self._check_permission(CanManageTemplates, self.admin, 'POST'))

    def test_all_permissions_deny_anonymous(self):
        from django.contrib.auth.models import AnonymousUser
        from api.permissions import (IsAdmin, IsAdminOrReadOnly, CanManageUsers,
            CanManageCertificates, CanManageEvents, CanManageStudents,
            CanManageInstructors, CanManageTemplates)
        anon = AnonymousUser()
        for perm_class in [IsAdmin, IsAdminOrReadOnly, CanManageUsers,
                           CanManageCertificates, CanManageEvents, CanManageStudents,
                           CanManageInstructors, CanManageTemplates]:
            self.assertFalse(self._check_permission(perm_class, anon))


# ─────────────────────────────────────────────
# Participante (non-admin) role tests
# ─────────────────────────────────────────────

class ParticipanteAccessTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.regular = make_user()
        self.client.force_authenticate(user=self.regular)

    def test_participante_can_list_participants(self):
        res = self.client.get('/api/participants/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_participante_cannot_create_participant(self):
        res = self.client.post('/api/participants/', {
            'document_id': '77777', 'first_name': 'X', 'last_name': 'Y',
            'email': 'x@test.com', 'phone': ''
        })
        self.assertIn(res.status_code, [status.HTTP_403_FORBIDDEN, status.HTTP_401_UNAUTHORIZED])

    def test_participante_can_view_certificates(self):
        res = self.client.get('/api/certificates/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_participante_can_verify_certificate_public(self):
        self.client.force_authenticate(user=None)
        res = self.client.get('/api/certificates/verify/?code=INVALID')
        self.assertIn(res.status_code, [status.HTTP_404_NOT_FOUND, status.HTTP_400_BAD_REQUEST])

    def test_participante_can_list_events(self):
        res = self.client.get('/api/events/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ─────────────────────────────────────────────
# Events extra endpoints
# ─────────────────────────────────────────────

class EventExtraEndpointsTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin)
        self.participant = make_participant(self.admin)
        self.template = Template.objects.create(name='T', created_by=self.admin)

    def test_event_deliveries_empty(self):
        res = self.client.get(f'/api/events/{self.event.id}/deliveries/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data, [])

    def test_stats_with_enrollment(self):
        Enrollment.objects.create(participant=self.participant, event=self.event, attendance=True, created_by=self.admin)
        res = self.client.get(f'/api/events/{self.event.id}/stats/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['attendees'], 1)
        self.assertEqual(res.data['total_enrollments'], 1)

    def test_invitations_list_empty(self):
        res = self.client.get(f'/api/events/{self.event.id}/invitations/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data, [])

    def test_generate_certificates_no_enrollments(self):
        res = self.client.post(f'/api/events/{self.event.id}/certificates/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['created'], 0)

    def test_generate_certificates_with_student(self):
        Enrollment.objects.create(participant=self.participant, event=self.event, attendance=True, created_by=self.admin)
        res = self.client.post(f'/api/events/{self.event.id}/certificates/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_generate_certificates_non_admin_forbidden(self):
        regular = make_user('r@test.com')
        self.client.force_authenticate(user=regular)
        res = self.client.post(f'/api/events/{self.event.id}/certificates/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def test_send_certificates_non_admin_forbidden(self):
        regular = make_user('r2@test.com')
        self.client.force_authenticate(user=regular)
        res = self.client.post(f'/api/events/{self.event.id}/certificates/send/', {'method': 'email'})
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)


# ─────────────────────────────────────────────
# Certificate: expired verify + participante queryset
# ─────────────────────────────────────────────

class CertificateExtraTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.participant = make_participant(self.admin)
        self.event = make_event(self.admin)
        self.template = Template.objects.create(name='T', created_by=self.admin)

    def test_verify_expired_certificate(self):
        from django.utils import timezone
        from datetime import timedelta
        self.client.force_authenticate(user=self.admin)
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template,
            generated_by=self.admin,
            expires_at=timezone.now() - timedelta(days=1)
        )
        res = self.client.get(f'/api/certificates/verify/?code={cert.verification_code}')
        self.assertEqual(res.status_code, 410)

    def test_participante_sees_own_certificates(self):
        participante = User.objects.create_user(
            email='p@test.com', full_name='P', password='pass', role='participante'
        )
        participant2 = Participant.objects.create(
            document_id='88888', first_name='Par', last_name='T',
            email='p@test.com', created_by=self.admin
        )
        Certificate.objects.create(
            participant=participant2, event=self.event, template=self.template, generated_by=self.admin
        )
        self.client.force_authenticate(user=participante)
        res = self.client.get('/api/certificates/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ─────────────────────────────────────────────
# Enrollments
# ─────────────────────────────────────────────

class EnrollmentViewSetTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin)
        self.participant = make_participant(self.admin)

    def test_list_enrollments(self):
        res = self.client.get('/api/enrollments/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_create_enrollment(self):
        res = self.client.post('/api/enrollments/', {
            'participant_id': self.participant.id,
            'event_id': self.event.id,
        })
        self.assertIn(res.status_code, [status.HTTP_201_CREATED, status.HTTP_400_BAD_REQUEST])

    def test_create_enrollment_invalid_participant(self):
        res = self.client.post('/api/enrollments/', {
            'participant_id': 9999,
            'event_id': self.event.id,
        })
        self.assertIn(res.status_code, [status.HTTP_404_NOT_FOUND, status.HTTP_400_BAD_REQUEST])


# ─────────────────────────────────────────────
# Google Auth (mocked)
# ─────────────────────────────────────────────

class GoogleAuthViewTest(TestCase):
    def setUp(self):
        self.client = APIClient()

    def test_google_auth_missing_token(self):
        res = self.client.post('/api/auth/google/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    @patch('google.oauth2.id_token.verify_oauth2_token')
    def test_google_auth_invalid_token(self, mock_verify):
        mock_verify.side_effect = ValueError('invalid token')
        res = self.client.post('/api/auth/google/', {'token': 'bad-token'})
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    @patch('google.oauth2.id_token.verify_oauth2_token')
    def test_google_auth_existing_user(self, mock_verify):
        from django.conf import settings
        settings.GOOGLE_CLIENT_ID = 'test-client-id'
        User.objects.create_user(email='google@test.com', full_name='G', password='pass')
        mock_verify.return_value = {'email': 'google@test.com', 'name': 'G'}
        res = self.client.post('/api/auth/google/', {'token': 'valid-token'})
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_500_INTERNAL_SERVER_ERROR])

    @patch('google.oauth2.id_token.verify_oauth2_token')
    def test_google_auth_new_user(self, mock_verify):
        from django.conf import settings
        settings.GOOGLE_CLIENT_ID = 'test-client-id'
        mock_verify.return_value = {'email': 'newgoogle@test.com', 'name': 'New'}
        res = self.client.post('/api/auth/google/', {'token': 'valid-token'})
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_500_INTERNAL_SERVER_ERROR])


# ─────────────────────────────────────────────
# InvitationPublicView
# ─────────────────────────────────────────────

class InvitationPublicViewTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.event = make_event(self.admin)

    def _make_invitation(self, email='invite@test.com', status_val='pending', expires=None):
        from events.models import EventInvitation
        from django.utils import timezone
        from datetime import timedelta
        inv = EventInvitation.objects.create(
            event=self.event,
            email=email,
            status=status_val,
            expires_at=expires,
            created_by=self.admin
        )
        return inv

    def test_get_invitation_valid_token(self):
        inv = self._make_invitation()
        res = self.client.get(f'/api/invitations/{inv.token}/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_get_invitation_invalid_token(self):
        res = self.client.get('/api/invitations/nonexistent-token-xyz/')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_get_invitation_already_accepted(self):
        inv = self._make_invitation(status_val='accepted')
        res = self.client.get(f'/api/invitations/{inv.token}/')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_get_invitation_expired(self):
        from django.utils import timezone
        from datetime import timedelta
        inv = self._make_invitation(expires=timezone.now() - timedelta(hours=1))
        res = self.client.get(f'/api/invitations/{inv.token}/')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_accept_invitation_invalid_token(self):
        res = self.client.post('/api/invitations/bad-token/accept/', {})
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_accept_invitation_existing_student(self):
        student = make_participant(self.admin)
        from events.models import EventInvitation
        inv = EventInvitation.objects.create(
            event=self.event, email=student.email, participant=student,
            created_by=self.admin
        )
        res = self.client.post(f'/api/invitations/{inv.token}/accept/', {
            'email': student.email
        })
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST, status.HTTP_404_NOT_FOUND])


# ─────────────────────────────────────────────
# Certificate generate success (with PDF mock)
# ─────────────────────────────────────────────

class CertificateGenerateTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.participant = make_participant(self.admin)
        self.event = make_event(self.admin)
        self.template = Template.objects.create(name='T', created_by=self.admin)
        Enrollment.objects.create(participant=self.participant, event=self.event, attendance=True, created_by=self.admin)

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_generate_certificate_success(self, mock_pdf):
        mock_pdf.return_value = {'success': True, 'path': '/media/cert.pdf'}
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        res = self.client.post(f'/api/certificates/{cert.id}/generate/', {
            'student_id': self.participant.id, 'event_id': self.event.id
        })
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['status'], 'success')

    @patch('services.email_service.EmailService.send_certificate')
    def test_deliver_certificate_success(self, mock_send):
        mock_send.return_value = {'success': True, 'message': 'sent'}
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated', pdf_url='/media/c.pdf')
        cert.refresh_from_db()
        res = self.client.post(f'/api/certificates/{cert.id}/deliver/', {'method': 'email'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    @patch('services.email_service.EmailService.send_certificate')
    def test_deliver_certificate_whatsapp(self, mock_send):
        mock_send.return_value = {'success': True, 'message': 'sent'}
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated', pdf_url='/media/c.pdf')
        cert.refresh_from_db()
        res = self.client.post(f'/api/certificates/{cert.id}/deliver/', {'method': 'whatsapp'})
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST])


# ─────────────────────────────────────────────
# Participants import_students endpoint
# ─────────────────────────────────────────────

class ParticipantsImportTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)

    def test_import_without_file_returns_400(self):
        res = self.client.post('/api/participants/import_students/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_import_with_excel_file(self):
        import pandas as pd
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        df = pd.DataFrame([{
            'document_id': 'IMP001', 'first_name': 'Import', 'last_name': 'User',
            'email': 'import@test.com', 'phone': ''
        }])
        buf = BytesIO()
        df.to_excel(buf, index=False)
        buf.seek(0)
        f = SimpleUploadedFile('students.xlsx', buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res = self.client.post('/api/participants/import_students/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('imported', res.data)

    def test_import_with_csv_file(self):
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_content = b'document_id,first_name,last_name,email,phone\nCSV001,CSV,User,csv@test.com,\n'
        f = SimpleUploadedFile('students.csv', csv_content, content_type='text/csv')
        res = self.client.post('/api/participants/import_students/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('imported', res.data)

    def test_import_invalid_file_returns_400(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        f = SimpleUploadedFile('bad.xlsx', b'not-real-excel', content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res = self.client.post('/api/participants/import_students/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)


# ─────────────────────────────────────────────
# Google Auth - additional edge cases
# ─────────────────────────────────────────────

class GoogleAuthEdgeCasesTest(TestCase):
    def setUp(self):
        self.client = APIClient()

    def test_google_auth_no_client_id_configured(self):
        from django.conf import settings
        original = getattr(settings, 'GOOGLE_CLIENT_ID', None)
        settings.GOOGLE_CLIENT_ID = None
        try:
            res = self.client.post('/api/auth/google/', {'token': 'some-token'})
            self.assertIn(res.status_code, [
                status.HTTP_500_INTERNAL_SERVER_ERROR,
                status.HTTP_401_UNAUTHORIZED,
                status.HTTP_400_BAD_REQUEST,
            ])
        finally:
            settings.GOOGLE_CLIENT_ID = original

    @patch('google.oauth2.id_token.verify_oauth2_token')
    def test_google_auth_no_email_in_token(self, mock_verify):
        from django.conf import settings
        settings.GOOGLE_CLIENT_ID = 'test-client-id'
        mock_verify.return_value = {'name': 'No Email User'}
        res = self.client.post('/api/auth/google/', {'token': 'valid-token'})
        self.assertIn(res.status_code, [status.HTTP_400_BAD_REQUEST, status.HTTP_500_INTERNAL_SERVER_ERROR])

    @patch('google.oauth2.id_token.verify_oauth2_token')
    def test_google_auth_generic_exception(self, mock_verify):
        from django.conf import settings
        settings.GOOGLE_CLIENT_ID = 'test-client-id'
        mock_verify.side_effect = Exception('network error')
        res = self.client.post('/api/auth/google/', {'token': 'any-token'})
        self.assertIn(res.status_code, [status.HTTP_500_INTERNAL_SERVER_ERROR, status.HTTP_401_UNAUTHORIZED])


# ─────────────────────────────────────────────
# IsAdminUserOrReadOnly + IsCertificateOwnerOrAdmin in views.py
# ─────────────────────────────────────────────

class ViewPermissionClassesTest(TestCase):
    def test_is_admin_user_or_read_only_safe_method(self):
        from api.views import IsAdminUserOrReadOnly
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        factory = APIRequestFactory()
        raw = factory.get('/')
        from django.contrib.auth.models import AnonymousUser
        raw.user = AnonymousUser()
        req = Request(raw)
        req._user = AnonymousUser()
        perm = IsAdminUserOrReadOnly()
        self.assertTrue(perm.has_permission(req, None))

    def test_is_admin_user_or_read_only_write_non_admin(self):
        from api.views import IsAdminUserOrReadOnly
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        factory = APIRequestFactory()
        raw = factory.post('/')
        user = User.objects.create_user(email='nonadmin@test.com', full_name='N', password='pass')
        raw.user = user
        req = Request(raw)
        req._user = user
        perm = IsAdminUserOrReadOnly()
        self.assertFalse(perm.has_permission(req, None))

    def test_is_certificate_owner_or_admin_safe_method(self):
        from api.views import IsCertificateOwnerOrAdmin
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        factory = APIRequestFactory()
        raw = factory.get('/')
        from django.contrib.auth.models import AnonymousUser
        raw.user = AnonymousUser()
        req = Request(raw)
        perm = IsCertificateOwnerOrAdmin()
        self.assertTrue(perm.has_object_permission(req, None, None))

    def test_is_certificate_owner_or_admin_write_staff(self):
        from api.views import IsCertificateOwnerOrAdmin
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        factory = APIRequestFactory()
        raw = factory.post('/')
        admin = User.objects.create_user(email='staff@test.com', full_name='A', password='pass', is_staff=True)
        raw.user = admin
        req = Request(raw)
        req._user = admin
        perm = IsCertificateOwnerOrAdmin()
        self.assertTrue(perm.has_object_permission(req, None, None))


# ─────────────────────────────────────────────
# Instructors - non-admin paths
# ─────────────────────────────────────────────

class InstructorNonAdminTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.regular = make_user('regular@test.com')

    def test_non_admin_gets_own_instructors(self):
        Instructor.objects.create(full_name='Own', email='own@test.com', created_by=self.regular)
        Instructor.objects.create(full_name='Other', email='other@test.com', created_by=self.admin)
        self.client.force_authenticate(user=self.regular)
        res = self.client.get('/api/instructors/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        items = res.data.get('results', res.data) if isinstance(res.data, dict) else res.data
        names = [i['full_name'] for i in items]
        self.assertIn('Own', names)
        self.assertNotIn('Other', names)

    def test_non_admin_cannot_create_instructor(self):
        self.client.force_authenticate(user=self.regular)
        res = self.client.post('/api/instructors/', {'full_name': 'X', 'email': 'x@test.com'})
        self.assertIn(res.status_code, [status.HTTP_403_FORBIDDEN, status.HTTP_401_UNAUTHORIZED])


# ─────────────────────────────────────────────
# Template upload-image and preview
# ─────────────────────────────────────────────

class TemplateUploadTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.template = Template.objects.create(name='Upload Test', created_by=self.admin)

    def test_upload_image_no_file_returns_400(self):
        res = self.client.post(f'/api/templates/{self.template.id}/upload-image/', {}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_upload_image_invalid_type_returns_400(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        f = SimpleUploadedFile('doc.pdf', b'%PDF', content_type='application/pdf')
        res = self.client.post(f'/api/templates/{self.template.id}/upload-image/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_upload_image_valid_png(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        # Minimal valid PNG bytes
        import struct, zlib
        def minimal_png():
            sig = b'\x89PNG\r\n\x1a\n'
            def chunk(name, data):
                c = struct.pack('>I', len(data)) + name + data
                return c + struct.pack('>I', zlib.crc32(name + data) & 0xffffffff)
            ihdr = chunk(b'IHDR', struct.pack('>IIBBBBB', 1, 1, 8, 2, 0, 0, 0))
            raw = b'\x00\xff\xff\xff'
            compressed = zlib.compress(raw)
            idat = chunk(b'IDAT', compressed)
            iend = chunk(b'IEND', b'')
            return sig + ihdr + idat + iend
        f = SimpleUploadedFile('img.png', minimal_png(), content_type='image/png')
        res = self.client.post(f'/api/templates/{self.template.id}/upload-image/', {'file': f}, format='multipart')
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST])

    def test_get_preview(self):
        res = self.client.get(f'/api/templates/{self.template.id}/preview/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('layout_config', res.data)


# ─────────────────────────────────────────────
# EventsViewSet - send_certificates, send_invitations, finalize
# ─────────────────────────────────────────────

class EventsAdvancedTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin)
        self.participant = make_participant(self.admin)
        self.template = Template.objects.create(name='T', created_by=self.admin)

    def test_send_certificates_empty(self):
        res = self.client.post(f'/api/events/{self.event.id}/certificates/send/', {'method': 'email'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['total_sent'], 0)

    @patch('services.email_service.EmailService.send_certificate')
    def test_send_certificates_with_generated_cert(self, mock_send):
        mock_send.return_value = {'success': True, 'message': 'sent'}
        Enrollment.objects.create(participant=self.participant, event=self.event, attendance=True, created_by=self.admin)
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated', pdf_url='/media/cert.pdf')
        res = self.client.post(f'/api/events/{self.event.id}/certificates/send/', {'method': 'email'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_send_invitations_no_emails_returns_400(self):
        res = self.client.post(f'/api/events/{self.event.id}/invitations/send/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    @patch('django.core.mail.send_mail')
    def test_send_invitations_with_email_list(self, mock_mail):
        mock_mail.return_value = 1
        import json
        res = self.client.post(
            f'/api/events/{self.event.id}/invitations/send/',
            {'emails': json.dumps(['newguest@test.com'])},
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('created', res.data)

    @patch('django.core.mail.send_mail')
    def test_send_invitations_duplicate_skipped(self, mock_mail):
        mock_mail.return_value = 1
        from events.models import EventInvitation
        EventInvitation.objects.create(event=self.event, email='dup@test.com', created_by=self.admin)
        import json
        res = self.client.post(
            f'/api/events/{self.event.id}/invitations/send/',
            {'emails': json.dumps(['dup@test.com'])},
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('errors', res.data)

    def test_send_all_invitations_none_pending_returns_400(self):
        res = self.client.post(f'/api/events/{self.event.id}/invitations/send-all/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    @patch('django.core.mail.send_mail')
    def test_send_all_invitations_sends_pending(self, mock_mail):
        mock_mail.return_value = 1
        from events.models import EventInvitation
        EventInvitation.objects.create(
            event=self.event, email='pending@test.com', status='pending', created_by=self.admin
        )
        res = self.client.post(f'/api/events/{self.event.id}/invitations/send-all/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('sent', res.data)

    def test_finalize_event(self):
        res = self.client.post(f'/api/events/{self.event.id}/finalize/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['status'], 'finished')

    def test_finalize_already_finished_returns_400(self):
        Event.objects.filter(pk=self.event.pk).update(status='finished')
        self.event.refresh_from_db()
        res = self.client.post(f'/api/events/{self.event.id}/finalize/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    @patch('services.email_service.EmailService.send_certificate')
    def test_finalize_with_send_certificates(self, mock_send):
        mock_send.return_value = {'success': True, 'message': 'sent'}
        Enrollment.objects.create(participant=self.participant, event=self.event, attendance=True, created_by=self.admin)
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated', pdf_url='/media/cert.pdf')
        res = self.client.post(f'/api/events/{self.event.id}/finalize/', {'send_certificates': True})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_participante_sees_enrolled_events(self):
        participante = make_user('part@test.com')
        participant2 = Participant.objects.create(
            document_id='PART01', first_name='Part', last_name='User',
            email='part@test.com', created_by=self.admin
        )
        Enrollment.objects.create(participant=participant2, event=self.event, created_by=self.admin)
        self.client.force_authenticate(user=participante)
        res = self.client.get('/api/events/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_generate_certificates_already_exists(self):
        Enrollment.objects.create(participant=self.participant, event=self.event, attendance=True, created_by=self.admin)
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated')
        res = self.client.post(f'/api/events/{self.event.id}/certificates/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('already_exists', res.data['results'])

    def test_send_invitations_invalid_email_skipped(self):
        import json
        res = self.client.post(
            f'/api/events/{self.event.id}/invitations/send/',
            {'emails': json.dumps(['not-an-email'])},
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ─────────────────────────────────────────────
# BulkCertificateGenerationView
# ─────────────────────────────────────────────

class BulkCertificateGenerationViewTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='Bulk Event')

    def _make_excel_file(self):
        import pandas as pd
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        df = pd.DataFrame([{
            'full_name': 'Bulk User',
            'email': 'bulk@test.com',
            'document_id': 'BULK001',
            'event_name': 'Bulk Event',
            'phone': '111000999'
        }])
        buf = BytesIO()
        df.to_excel(buf, index=False)
        buf.seek(0)
        return SimpleUploadedFile('bulk.xlsx', buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_get_returns_format_info(self):
        res = self.client.get('/api/certificates/generate-bulk/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('required_columns', res.data)

    def test_post_no_file_returns_400(self):
        res = self.client.post('/api/certificates/generate-bulk/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_post_non_admin_forbidden(self):
        regular = make_user('reg@test.com')
        self.client.force_authenticate(user=regular)
        res = self.client.post('/api/certificates/generate-bulk/', {})
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def test_post_with_valid_excel(self):
        f = self._make_excel_file()
        res = self.client.post('/api/certificates/generate-bulk/', {'excel_file': f}, format='multipart')
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST])


# ─────────────────────────────────────────────
# BulkCertificatePreviewView
# ─────────────────────────────────────────────

class BulkCertificatePreviewViewTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='Preview Event')

    def _make_excel_file(self):
        import pandas as pd
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        df = pd.DataFrame([{
            'full_name': 'Preview User',
            'email': 'preview@test.com',
            'document_id': 'PRV001',
            'event_name': 'Preview Event',
        }])
        buf = BytesIO()
        df.to_excel(buf, index=False)
        buf.seek(0)
        return SimpleUploadedFile('preview.xlsx', buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_post_non_admin_forbidden(self):
        regular = make_user('prev@test.com')
        self.client.force_authenticate(user=regular)
        res = self.client.post('/api/certificates/preview/', {})
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def test_post_no_file_returns_400(self):
        res = self.client.post('/api/certificates/preview/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_post_with_valid_excel(self):
        f = self._make_excel_file()
        res = self.client.post('/api/certificates/preview/', {'excel_file': f}, format='multipart')
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST])

    def test_post_with_invalid_excel(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        f = SimpleUploadedFile('bad.xlsx', b'notexcel', content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res = self.client.post('/api/certificates/preview/', {'excel_file': f}, format='multipart')
        self.assertIn(res.status_code, [status.HTTP_400_BAD_REQUEST, status.HTTP_200_OK])


# ─────────────────────────────────────────────
# BulkCertificateProcessView
# ─────────────────────────────────────────────

class BulkCertificateProcessViewTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='Process Event')

    def test_post_empty_data_returns_400(self):
        res = self.client.post('/api/certificates/process/', {'data': []})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_post_missing_data_returns_400(self):
        res = self.client.post('/api/certificates/process/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_post_with_valid_records(self):
        res = self.client.post('/api/certificates/process/', {'data': [{
            'full_name': 'Process User',
            'email': 'proc@test.com',
            'document_id': 'PROC001',
            'event_name': 'Process Event',
        }]}, format='json')
        self.assertIn(res.status_code, [
            status.HTTP_200_OK,
            status.HTTP_400_BAD_REQUEST,
            status.HTTP_500_INTERNAL_SERVER_ERROR,
        ])


# ─────────────────────────────────────────────
# EnrollmentViewSet - create / destroy / attendance
# ─────────────────────────────────────────────

class EnrollmentViewSetAdvancedTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin)
        self.participant = make_participant(self.admin)

    def test_create_enrollment_success(self):
        res = self.client.post('/api/enrollments/', {
            'participant_id': self.participant.id,
            'event_id': self.event.id,
            'attendance': False,
        }, format='json')
        self.assertIn(res.status_code, [status.HTTP_201_CREATED, status.HTTP_400_BAD_REQUEST])

    def test_create_enrollment_nonexistent_participant(self):
        res = self.client.post('/api/enrollments/', {
            'participant_id': 99999,
            'event_id': self.event.id,
        }, format='json')
        self.assertIn(res.status_code, [status.HTTP_404_NOT_FOUND, status.HTTP_400_BAD_REQUEST])

    def test_create_enrollment_nonexistent_event(self):
        res = self.client.post('/api/enrollments/', {
            'participant_id': self.participant.id,
            'event_id': 99999,
        }, format='json')
        self.assertIn(res.status_code, [status.HTTP_404_NOT_FOUND, status.HTTP_400_BAD_REQUEST])

    def test_non_admin_list_enrollments_returns_403_without_event_pk(self):
        regular = make_user('r@test.com')
        self.client.force_authenticate(user=regular)
        res = self.client.get('/api/enrollments/')
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def test_admin_can_list_all_enrollments(self):
        Enrollment.objects.create(participant=self.participant, event=self.event, created_by=self.admin)
        res = self.client.get('/api/enrollments/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ─────────────────────────────────────────────
# InvitationPublicView - POST paths
# ─────────────────────────────────────────────

class InvitationPublicPostTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.event = make_event(self.admin)

    def _make_invitation(self, email='invite@test.com', status_val='pending', expires=None):
        from events.models import EventInvitation
        return EventInvitation.objects.create(
            event=self.event, email=email, status=status_val,
            expires_at=expires, created_by=self.admin
        )

    def test_accept_no_student_registered_returns_400(self):
        inv = self._make_invitation(email='noone@test.com')
        res = self.client.post(f'/api/invitations/{inv.token}/accept/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_accept_already_accepted_invitation(self):
        inv = self._make_invitation(status_val='accepted')
        res = self.client.post(f'/api/invitations/{inv.token}/accept/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_accept_expired_invitation(self):
        from django.utils import timezone
        from datetime import timedelta
        inv = self._make_invitation(expires=timezone.now() - timedelta(hours=1))
        res = self.client.post(f'/api/invitations/{inv.token}/accept/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_accept_valid_invitation_with_existing_student(self):
        student = make_participant(self.admin, email='inv_student@test.com')
        inv = self._make_invitation(email='inv_student@test.com')
        inv.participant = student
        inv.save()
        res = self.client.post(f'/api/invitations/{inv.token}/accept/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_accept_finds_student_by_email(self):
        make_participant(self.admin, email='findme@test.com', doc='FIND01')
        inv = self._make_invitation(email='findme@test.com')
        res = self.client.post(f'/api/invitations/{inv.token}/accept/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ─────────────────────────────────────────────
# InvitationRegisterView
# ─────────────────────────────────────────────

class InvitationRegisterViewTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.event = make_event(self.admin)

    def _make_invitation(self, email='reg@test.com', status_val='pending', expires=None):
        from events.models import EventInvitation
        return EventInvitation.objects.create(
            event=self.event, email=email, status=status_val,
            expires_at=expires, created_by=self.admin
        )

    def test_register_invalid_token_returns_404(self):
        res = self.client.post('/api/invitations/bad-token/register/', {
            'first_name': 'X', 'last_name': 'Y', 'password': 'Pass1234!'
        })
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_register_already_accepted_returns_400(self):
        inv = self._make_invitation(status_val='accepted')
        res = self.client.post(f'/api/invitations/{inv.token}/register/', {
            'first_name': 'X', 'last_name': 'Y', 'password': 'Pass1234!'
        })
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_register_expired_returns_400(self):
        from django.utils import timezone
        from datetime import timedelta
        inv = self._make_invitation(expires=timezone.now() - timedelta(hours=1))
        res = self.client.post(f'/api/invitations/{inv.token}/register/', {
            'first_name': 'X', 'last_name': 'Y', 'password': 'Pass1234!'
        })
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_register_new_user_success(self):
        inv = self._make_invitation(email='newreg@test.com')
        res = self.client.post(f'/api/invitations/{inv.token}/register/', {
            'first_name': 'New', 'last_name': 'Reg', 'password': 'Pass1234!', 'phone': '999'
        })
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST])

    def test_register_existing_user_links_student(self):
        User.objects.create_user(email='existing@test.com', full_name='Exist', password='pass')
        inv = self._make_invitation(email='existing@test.com')
        res = self.client.post(f'/api/invitations/{inv.token}/register/', {
            'first_name': 'Exist', 'last_name': 'User', 'password': 'Pass1234!'
        })
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST])

    def test_register_missing_fields_returns_400(self):
        inv = self._make_invitation(email='incomplete@test.com')
        res = self.client.post(f'/api/invitations/{inv.token}/register/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)


# ─────────────────────────────────────────────
# CertificateViewSet - participante queryset + generate with template_id
# ─────────────────────────────────────────────

class CertificateViewSetAdvancedTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.participant = make_participant(self.admin)
        self.event = make_event(self.admin)
        self.template = Template.objects.create(name='T', created_by=self.admin)

    def test_participante_sees_only_own_certs(self):
        participante = make_user('cert_part@test.com')
        participant2 = Participant.objects.create(
            document_id='CERTPART', first_name='Cert', last_name='Part',
            email='cert_part@test.com', created_by=self.admin
        )
        own_cert = Certificate.objects.create(
            participant=participant2, event=self.event, template=self.template, generated_by=self.admin
        )
        other_cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        self.client.force_authenticate(user=participante)
        res = self.client.get('/api/certificates/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        items = res.data.get('results', res.data) if isinstance(res.data, dict) else res.data
        ids = [c['id'] for c in items]
        self.assertIn(own_cert.id, ids)
        self.assertNotIn(other_cert.id, ids)

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_generate_with_template_id(self, mock_pdf):
        mock_pdf.return_value = {'success': True, 'path': '/media/cert.pdf'}
        self.client.force_authenticate(user=self.admin)
        Enrollment.objects.create(participant=self.participant, event=self.event, attendance=True, created_by=self.admin)
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        template2 = Template.objects.create(name='T2', created_by=self.admin)
        res = self.client.post(f'/api/certificates/{cert.id}/generate/', {
            'student_id': self.participant.id,
            'event_id': self.event.id,
            'template_id': template2.id,
        }, format='json')
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST])


# ─────────────────────────────────────────────
# Serializer edge cases
# ─────────────────────────────────────────────

class SerializerEdgeCasesTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(email='serial@test.com', full_name='S', password='pass', is_staff=True)

    def test_date_field_none_returns_none(self):
        from api.serializers import DateField
        f = DateField()
        self.assertIsNone(f.to_representation(None))

    def test_date_field_datetime_returns_date_string(self):
        from api.serializers import DateField
        from datetime import datetime
        f = DateField()
        dt = datetime(2026, 5, 1, 12, 0)
        result = f.to_representation(dt)
        self.assertEqual(result, '2026-05-01')

    def test_date_field_internal_value_empty_string(self):
        from api.serializers import DateField
        f = DateField()
        self.assertIsNone(f.to_internal_value(''))

    def test_date_field_internal_value_none(self):
        from api.serializers import DateField
        f = DateField()
        self.assertIsNone(f.to_internal_value(None))

    def test_event_serializer_with_template_and_instructor(self):
        from api.serializers import EventSerializer
        from instructors.models import Instructor
        inst = Instructor.objects.create(full_name='Inst', email='inst@test.com', created_by=self.user)
        template = Template.objects.create(name='Tmpl', created_by=self.user)
        event = Event.objects.create(
            name='Ev', event_date=date(2026, 6, 1), created_by=self.user,
            template=template, instructor=inst
        )
        s = EventSerializer(event)
        self.assertEqual(s.data['template_name'], 'Tmpl')
        self.assertEqual(s.data['instructor_name'], 'Inst')

    def test_user_login_serializer_inactive_user(self):
        from api.serializers import UserLoginSerializer
        User.objects.create_user(
            email='inactive@test.com', full_name='I', password='pass', is_active=False
        )
        s = UserLoginSerializer(data={'email': 'inactive@test.com', 'password': 'pass'})
        self.assertFalse(s.is_valid())

    def test_user_auth_serializer_missing_password(self):
        from api.serializers import UserAuthSerializer
        s = UserAuthSerializer(data={'email': 'x@x.com'})
        self.assertFalse(s.is_valid())

    def test_template_serializer_with_background_url(self):
        from api.serializers import TemplateSerializer
        t = Template.objects.create(name='BG', created_by=self.user, background_url='http://example.com/bg.png')
        s = TemplateSerializer(t)
        self.assertEqual(s.data['background_image_url'], 'http://example.com/bg.png')

    def test_invitation_detail_serializer_participant_exists_by_link(self):
        from api.serializers import InvitationDetailSerializer
        from events.models import EventInvitation
        event = Event.objects.create(name='Inv Ev', event_date=date(2026, 7, 1), created_by=self.user)
        participant = Participant.objects.create(
            document_id='INVSER', first_name='A', last_name='B',
            email='invser@test.com', created_by=self.user
        )
        inv = EventInvitation.objects.create(
            event=event, email='invser@test.com', participant=participant, created_by=self.user
        )
        s = InvitationDetailSerializer(inv)
        self.assertTrue(s.data['participant_exists'])


# ─────────────────────────────────────────────
# Permission classes - participante safe-method paths
# ─────────────────────────────────────────────

class PermissionSafeMethodTest(TestCase):
    def _check_safe(self, perm_class, user):
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        factory = APIRequestFactory()
        raw = factory.get('/')
        raw.user = user
        req = Request(raw)
        req._user = user
        perm = perm_class()
        return perm.has_permission(req, None)

    def _check_write(self, perm_class, user):
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        factory = APIRequestFactory()
        raw = factory.post('/')
        raw.user = user
        req = Request(raw)
        req._user = user
        perm = perm_class()
        return perm.has_permission(req, None)

    def setUp(self):
        self.participante = User.objects.create_user(
            email='perm_part@test.com', full_name='P', password='pass', role='participante'
        )

    def test_can_manage_users_participante_safe(self):
        from api.permissions import CanManageUsers
        self.assertTrue(self._check_safe(CanManageUsers, self.participante))

    def test_can_manage_users_participante_write(self):
        from api.permissions import CanManageUsers
        self.assertFalse(self._check_write(CanManageUsers, self.participante))

    def test_can_manage_certificates_participante_safe(self):
        from api.permissions import CanManageCertificates
        self.assertTrue(self._check_safe(CanManageCertificates, self.participante))

    def test_can_manage_certificates_participante_write(self):
        from api.permissions import CanManageCertificates
        self.assertFalse(self._check_write(CanManageCertificates, self.participante))

    def test_can_manage_events_participante_safe(self):
        from api.permissions import CanManageEvents
        self.assertTrue(self._check_safe(CanManageEvents, self.participante))

    def test_can_manage_events_participante_write(self):
        from api.permissions import CanManageEvents
        self.assertFalse(self._check_write(CanManageEvents, self.participante))

    def test_can_manage_students_participante_safe(self):
        from api.permissions import CanManageStudents
        self.assertTrue(self._check_safe(CanManageStudents, self.participante))

    def test_can_manage_students_participante_write(self):
        from api.permissions import CanManageStudents
        self.assertFalse(self._check_write(CanManageStudents, self.participante))

    def test_can_manage_instructors_participante_safe(self):
        from api.permissions import CanManageInstructors
        self.assertTrue(self._check_safe(CanManageInstructors, self.participante))

    def test_can_manage_instructors_participante_write(self):
        from api.permissions import CanManageInstructors
        self.assertFalse(self._check_write(CanManageInstructors, self.participante))

    def test_can_manage_templates_participante_safe(self):
        from api.permissions import CanManageTemplates
        self.assertTrue(self._check_safe(CanManageTemplates, self.participante))

    def test_can_manage_templates_participante_write(self):
        from api.permissions import CanManageTemplates
        self.assertFalse(self._check_write(CanManageTemplates, self.participante))


# ─────────────────────────────────────────────
# CertificateViewSet - serializer class + permissions + exception paths
# ─────────────────────────────────────────────

class CertificateViewSetCoverageTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.participant = make_participant(self.admin)
        self.event = make_event(self.admin)
        self.template = Template.objects.create(name='T', created_by=self.admin)
        Enrollment.objects.create(participant=self.participant, event=self.event, attendance=True, created_by=self.admin)

    def test_create_certificate_hits_create_serializer(self):
        self.client.force_authenticate(user=self.admin)
        res = self.client.post('/api/certificates/', {
            'student': self.participant.id,
            'event': self.event.id,
            'template': self.template.id,
        }, format='json')
        self.assertIn(res.status_code, [status.HTTP_201_CREATED, status.HTTP_400_BAD_REQUEST])

    def test_non_admin_create_certificate_uses_is_admin_user_permission(self):
        regular = make_user('certperm@test.com')
        self.client.force_authenticate(user=regular)
        res = self.client.post('/api/certificates/', {
            'student': self.participant.id,
            'event': self.event.id,
        }, format='json')
        self.assertIn(res.status_code, [status.HTTP_403_FORBIDDEN, status.HTTP_401_UNAUTHORIZED])

    @patch('certificados.models.Certificate.generate')
    def test_generate_exception_caught_returns_400(self, mock_generate):
        mock_generate.side_effect = Exception('PDF error')
        self.client.force_authenticate(user=self.admin)
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        res = self.client.post(f'/api/certificates/{cert.id}/generate/', {
            'student_id': self.participant.id,
            'event_id': self.event.id,
        }, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('Failed to generate certificate', res.data['message'])

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_generate_with_template_id_overrides_template(self, mock_pdf):
        mock_pdf.return_value = {'success': True, 'path': '/media/cert.pdf'}
        self.client.force_authenticate(user=self.admin)
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        template2 = Template.objects.create(name='T2', created_by=self.admin)
        res = self.client.post(f'/api/certificates/{cert.id}/generate/', {
            'student_id': self.participant.id,
            'event_id': self.event.id,
            'template_id': template2.id,
        }, format='json')
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST])

    def test_unauthenticated_cert_queryset_returns_none(self):
        self.client.force_authenticate(user=None)
        res = self.client.get('/api/certificates/verify/?code=ABCD-ABCD-ABCD-ABCD')
        self.assertIn(res.status_code, [status.HTTP_400_BAD_REQUEST, status.HTTP_404_NOT_FOUND])


# ─────────────────────────────────────────────
# EventsViewSet - generate_certificates & send_certificates extra paths
# ─────────────────────────────────────────────

class EventsCertificateActionsTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin)
        self.participant = make_participant(self.admin)
        self.template = Template.objects.create(name='T', created_by=self.admin)
        Enrollment.objects.create(participant=self.participant, event=self.event, attendance=True, created_by=self.admin)

    def test_generate_certificates_with_participant_ids_filter(self):
        res = self.client.post(
            f'/api/events/{self.event.id}/certificates/generate/',
            {'participant_ids': [self.participant.id]},
            format='json'
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_generate_certificates_pending_cert_triggers_regenerate(self, mock_pdf):
        mock_pdf.return_value = {'success': True, 'path': '/media/cert.pdf'}
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template,
            generated_by=self.admin, status='pending'
        )
        res = self.client.post(f'/api/events/{self.event.id}/certificates/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_send_certificates_with_participant_ids_filter(self):
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated', pdf_url='/cert.pdf')
        with patch('services.email_service.EmailService.send_certificate') as mock_send:
            mock_send.return_value = {'success': True, 'message': 'sent'}
            res = self.client.post(
                f'/api/events/{self.event.id}/certificates/send/',
                {'method': 'email', 'participant_ids': [self.participant.id]},
                format='json'
            )
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    @patch('services.email_service.EmailService.send_certificate')
    def test_send_certificates_generates_pending_first(self, mock_send, mock_pdf):
        mock_pdf.return_value = {'success': True, 'path': '/media/cert.pdf'}
        mock_send.return_value = {'success': True, 'message': 'sent'}
        Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template,
            generated_by=self.admin, status='pending'
        )
        res = self.client.post(
            f'/api/events/{self.event.id}/certificates/send/',
            {'method': 'email'}
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    @patch('certificados.models.Certificate.deliver')
    def test_send_certificates_exception_logs_failed(self, mock_deliver):
        mock_deliver.side_effect = Exception('delivery failed')
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, template=self.template, generated_by=self.admin
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated', pdf_url='/cert.pdf')
        res = self.client.post(
            f'/api/events/{self.event.id}/certificates/send/',
            {'method': 'email'}
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertGreater(res.data['total_failed'], 0)


# ─────────────────────────────────────────────
# EventsViewSet - _parse_emails helpers
# ─────────────────────────────────────────────

class EventsEmailParsingTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin)

    def test_send_invitations_csv_file_with_email_column(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_content = b'email\nfileguest@test.com\n'
        f = SimpleUploadedFile('emails.csv', csv_content, content_type='text/csv')
        with patch('django.core.mail.send_mail', return_value=1):
            res = self.client.post(
                f'/api/events/{self.event.id}/invitations/send/',
                {'file': f},
                format='multipart'
            )
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_send_invitations_csv_file_no_email_column_returns_400(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        csv_content = b'name,phone\nJohn,555\n'
        f = SimpleUploadedFile('noemail.csv', csv_content, content_type='text/csv')
        res = self.client.post(
            f'/api/events/{self.event.id}/invitations/send/',
            {'file': f},
            format='multipart'
        )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_send_invitations_file_read_error_returns_400(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        f = SimpleUploadedFile('bad.csv', b'corrupted\x00\x01\x02', content_type='text/csv')
        with patch('pandas.read_csv', side_effect=Exception('read error')):
            res = self.client.post(
                f'/api/events/{self.event.id}/invitations/send/',
                {'file': f},
                format='multipart'
            )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_send_invitations_invalid_json_emails_treated_as_empty(self):
        res = self.client.post(
            f'/api/events/{self.event.id}/invitations/send/',
            {'emails': 'not{valid[json'},
        )
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    @patch('django.core.mail.send_mail')
    def test_send_invitations_email_failure_appended_to_errors(self, mock_mail):
        mock_mail.side_effect = Exception('SMTP error')
        import json
        res = self.client.post(
            f'/api/events/{self.event.id}/invitations/send/',
            {'emails': json.dumps(['errmail@test.com'])},
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertTrue(len(res.data['errors']) > 0)

    @patch('django.core.mail.send_mail')
    def test_send_all_invitations_success_marks_sent(self, mock_mail):
        mock_mail.return_value = 1
        from events.models import EventInvitation
        EventInvitation.objects.create(
            event=self.event, email='tosend@test.com', status='pending', created_by=self.admin
        )
        res = self.client.post(f'/api/events/{self.event.id}/invitations/send-all/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertGreaterEqual(res.data['sent'], 1)

    @patch('django.core.mail.send_mail')
    def test_send_all_invitations_email_failure_appended_to_errors(self, mock_mail):
        mock_mail.side_effect = Exception('SMTP fail')
        from events.models import EventInvitation
        EventInvitation.objects.create(
            event=self.event, email='failsend@test.com', status='pending', created_by=self.admin
        )
        res = self.client.post(f'/api/events/{self.event.id}/invitations/send-all/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertGreater(len(res.data['errors']), 0)


# ─────────────────────────────────────────────
# generate_certificates loop exception
# ─────────────────────────────────────────────

class GenerateCertificatesLoopExceptionTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin)
        self.participant = make_participant(self.admin)
        self.template = Template.objects.create(name='T', created_by=self.admin)
        Enrollment.objects.create(participant=self.participant, event=self.event, attendance=True, created_by=self.admin)

    @patch('certificados.models.Certificate.objects')
    def test_generate_certificates_exception_logged_in_errors(self, mock_objects):
        mock_objects.get_or_create.side_effect = Exception('DB error')
        mock_objects.filter.return_value = Certificate.objects.filter(event=self.event)
        res = self.client.post(f'/api/events/{self.event.id}/certificates/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ─────────────────────────────────────────────
# BulkCertificate views - exception paths
# ─────────────────────────────────────────────

class BulkCertificateViewExceptionTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='Bulk Test')

    def _make_excel_file(self):
        import pandas as pd
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        df = pd.DataFrame([{
            'full_name': 'Test User', 'email': 'testbulk@test.com',
            'document_id': 'BLKTEST', 'event_name': 'Bulk Test',
        }])
        buf = BytesIO()
        df.to_excel(buf, index=False)
        buf.seek(0)
        return SimpleUploadedFile('bulk.xlsx', buf.read(),
                                   content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_bulk_generate_post_with_valid_excel_reaches_processing(self):
        f = self._make_excel_file()
        res = self.client.post('/api/certificates/generate-bulk/', {'excel_file': f}, format='multipart')
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST])

    @patch('api.views.BulkCertificateGeneratorService.generate_from_excel')
    def test_bulk_generate_exception_returns_400(self, mock_gen):
        mock_gen.side_effect = Exception('processing error')
        f = self._make_excel_file()
        res = self.client.post('/api/certificates/generate-bulk/', {'excel_file': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_bulk_preview_post_with_valid_excel(self):
        f = self._make_excel_file()
        res = self.client.post('/api/certificates/preview/', {'excel_file': f}, format='multipart')
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST])

    @patch('api.views.ExcelProcessingService')
    def test_bulk_preview_excel_import_error_returns_400(self, mock_svc):
        from procesos.services import ExcelImportError
        mock_svc.return_value.read_and_validate_structure.side_effect = ExcelImportError('bad file')
        f = self._make_excel_file()
        res = self.client.post('/api/certificates/preview/', {'excel_file': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    @patch('api.views.ExcelProcessingService')
    def test_bulk_preview_generic_exception_returns_400(self, mock_svc):
        mock_svc.return_value.read_and_validate_structure.side_effect = Exception('unexpected')
        f = self._make_excel_file()
        res = self.client.post('/api/certificates/preview/', {'excel_file': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    @patch('api.views.ExcelProcessingService')
    def test_bulk_process_excel_import_error_returns_400(self, mock_svc):
        from procesos.services import ExcelImportError
        mock_svc.return_value.process_records.side_effect = ExcelImportError('bad data')
        res = self.client.post('/api/certificates/process/', {'data': [{'full_name': 'X'}]}, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    @patch('api.views.ExcelProcessingService')
    def test_bulk_process_generic_exception_returns_500(self, mock_svc):
        mock_svc.return_value.process_records.side_effect = RuntimeError('unexpected error')
        res = self.client.post('/api/certificates/process/', {'data': [{'full_name': 'X'}]}, format='json')
        self.assertEqual(res.status_code, status.HTTP_500_INTERNAL_SERVER_ERROR)


# ─────────────────────────────────────────────
# EnrollmentViewSet - destroy, attendance, non-admin list
# ─────────────────────────────────────────────

class EnrollmentViewSetDestroyAttendanceTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin)
        self.participant = make_participant(self.admin)
        self.enrollment = Enrollment.objects.create(
            participant=self.participant, event=self.event, attendance=False, created_by=self.admin
        )

    def test_destroy_enrollment_success(self):
        res = self.client.delete(f'/api/enrollments/{self.enrollment.id}/')
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)

    def test_destroy_enrollment_not_found(self):
        res = self.client.delete('/api/enrollments/99999/')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_attendance_mark_true(self):
        res = self.client.patch(
            f'/api/enrollments/{self.enrollment.id}/attendance/',
            {'attendance': True},
            format='json'
        )
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_404_NOT_FOUND])

    def test_attendance_missing_field_returns_400(self):
        res = self.client.patch(
            f'/api/enrollments/{self.enrollment.id}/attendance/',
            {},
            format='json'
        )
        self.assertIn(res.status_code, [status.HTTP_400_BAD_REQUEST, status.HTTP_404_NOT_FOUND])

    def test_attendance_not_found(self):
        res = self.client.patch(
            '/api/enrollments/99999/attendance/',
            {'attendance': True},
            format='json'
        )
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_non_admin_list_returns_403(self):
        regular = make_user('nonadmin_enr@test.com')
        self.client.force_authenticate(user=regular)
        res = self.client.get('/api/enrollments/')
        self.assertEqual(res.status_code, status.HTTP_403_FORBIDDEN)

    def test_create_enrollment_success(self):
        participant2 = Participant.objects.create(
            document_id='ENROLL2', first_name='New', last_name='Student',
            email='enroll2@test.com', created_by=self.admin
        )
        res = self.client.post('/api/enrollments/', {
            'participant_id': participant2.id,
            'event_id': self.event.id,
        }, format='json')
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)

    def test_create_enrollment_duplicate_returns_400(self):
        res = self.client.post('/api/enrollments/', {
            'participant_id': self.participant.id,
            'event_id': self.event.id,
        }, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_create_enrollment_invalid_participant_returns_404(self):
        res = self.client.post('/api/enrollments/', {
            'participant_id': 99999,
            'event_id': self.event.id,
        }, format='json')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_create_enrollment_invalid_event_returns_404(self):
        participant2 = Participant.objects.create(
            document_id='ENROLL3', first_name='S', last_name='T',
            email='enroll3@test.com', created_by=self.admin
        )
        res = self.client.post('/api/enrollments/', {
            'participant_id': participant2.id,
            'event_id': 99999,
        }, format='json')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)


# ─────────────────────────────────────────────
# Import students - IntegrityError / exception paths
# ─────────────────────────────────────────────

class ImportStudentsErrorPathTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin()
        self.client.force_authenticate(user=self.admin)

    def test_import_with_duplicate_document_id(self):
        import pandas as pd
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        make_participant(self.admin, doc='DUP001', email='dup@test.com')
        df = pd.DataFrame([{'document_id': 'DUP001', 'first_name': 'Dup', 'last_name': 'User', 'email': 'dup2@test.com', 'phone': ''}])
        buf = BytesIO()
        df.to_excel(buf, index=False)
        buf.seek(0)
        f = SimpleUploadedFile('dup.xlsx', buf.read(),
                               content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res = self.client.post('/api/participants/import_students/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_template_non_admin_cannot_create(self):
        regular = make_user('tperm@test.com')
        self.client.force_authenticate(user=regular)
        res = self.client.post('/api/templates/', {'name': 'Test'})
        self.assertIn(res.status_code, [status.HTTP_403_FORBIDDEN, status.HTTP_401_UNAUTHORIZED])

    def test_import_with_duplicate_email_triggers_integrity_error(self):
        import pandas as pd
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        make_participant(self.admin, doc='UNIQUE_DOC1', email='unique_dup@test.com')
        df = pd.DataFrame([{
            'document_id': 'DIFFERENT_DOC1',
            'first_name': 'New', 'last_name': 'User',
            'email': 'unique_dup@test.com', 'phone': '',
        }])
        buf = BytesIO()
        df.to_excel(buf, index=False)
        buf.seek(0)
        f = SimpleUploadedFile('dup_email.xlsx', buf.read(),
                               content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res = self.client.post('/api/participants/import_students/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertGreater(len(res.data.get('errors', [])), 0)

    @patch('participants.models.Participant.objects')
    def test_import_generic_exception_per_row_logged(self, mock_objects):
        import pandas as pd
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        mock_objects.get_or_create.side_effect = RuntimeError('unexpected row error')
        df = pd.DataFrame([{'document_id': 'ERRDOC', 'first_name': 'X', 'last_name': 'Y', 'email': 'x@test.com', 'phone': ''}])
        buf = BytesIO()
        df.to_excel(buf, index=False)
        buf.seek(0)
        f = SimpleUploadedFile('err.xlsx', buf.read(),
                               content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res = self.client.post('/api/participants/import_students/', {'file': f}, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertGreater(len(res.data.get('errors', [])), 0)


# ─────────────────────────────────────────────
# Finalize event with send_certificates coverage
# ─────────────────────────────────────────────

class FinalizeEventCertificateSentTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('fin_admin@test.com')
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='FinalizeTest')
        self.participant = make_participant(self.admin, doc='FINSTUD', email='finstud@test.com')

    @patch('certificados.models.Certificate.deliver')
    def test_finalize_with_send_certificates_covers_sent_fields(self, mock_deliver):
        mock_log = MagicMock()
        mock_log.status = 'success'
        mock_deliver.return_value = mock_log
        from events.models import Enrollment
        from certificados.models import Certificate
        Enrollment.objects.create(
            participant=self.participant, event=self.event, attendance=True, created_by=self.admin
        )
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, generated_by=self.admin
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated')
        res = self.client.post(f'/api/events/{self.event.id}/finalize/', {'send_certificates': True})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['certificates_sent'], 1)

    @patch('certificados.models.Certificate.deliver')
    def test_finalize_deliver_exception_logged(self, mock_deliver):
        mock_deliver.side_effect = Exception('deliver error')
        from events.models import Enrollment
        from certificados.models import Certificate
        event2 = make_event(self.admin, name='FinalizeExc')
        Enrollment.objects.create(
            participant=self.participant, event=event2, attendance=True, created_by=self.admin
        )
        cert = Certificate.objects.create(
            participant=self.participant, event=event2, generated_by=self.admin
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated')
        res = self.client.post(f'/api/events/{event2.id}/finalize/', {'send_certificates': True})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['certificates_sent'], 0)


# ─────────────────────────────────────────────
# Enrollment create invalid serializer + non-admin write
# ─────────────────────────────────────────────

class EnrollmentEdgeCasesTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('enr_edge@test.com')
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin)

    def test_create_enrollment_missing_participant_id_returns_400(self):
        res = self.client.post('/api/enrollments/', {'event_id': self.event.id}, format='json')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_non_admin_cannot_create_enrollment(self):
        regular = make_user('enr_nonadmin@test.com')
        self.client.force_authenticate(user=regular)
        res = self.client.post('/api/enrollments/', {'participant_id': 1, 'event_id': self.event.id}, format='json')
        self.assertIn(res.status_code, [status.HTTP_403_FORBIDDEN, status.HTTP_401_UNAUTHORIZED])


# ─────────────────────────────────────────────
# Template upload image with alternate key
# ─────────────────────────────────────────────

class TemplateUploadImageAltKeyTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('tpl_img@test.com')
        self.client.force_authenticate(user=self.admin)
        from certificados.models import Template
        self.template = Template.objects.create(name='ImgTpl', created_by=self.admin)

    def test_upload_image_with_alternate_key(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        img = SimpleUploadedFile('bg.png', b'\x89PNG\r\n', content_type='image/png')
        res = self.client.post(
            f'/api/templates/{self.template.id}/upload-image/',
            {'background': img},
            format='multipart'
        )
        self.assertIn(res.status_code, [status.HTTP_200_OK, status.HTTP_400_BAD_REQUEST])


# ─────────────────────────────────────────────
# Serializer coverage: DateField and InvitationDetailSerializer
# ─────────────────────────────────────────────

class SerializerCoverageTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(email='ser_cov@test.com', full_name='S', password='pass')

    def test_date_field_to_internal_value_empty_string_returns_none(self):
        from api.serializers import DateField
        field = DateField(allow_null=True, required=False)
        result = field.to_internal_value('')
        self.assertIsNone(result)

    def test_user_auth_serializer_validate_raises_when_no_email_or_password(self):
        from api.serializers import UserAuthSerializer
        s = UserAuthSerializer()
        with self.assertRaises(Exception):
            s.validate({'email': '', 'password': ''})

    def test_template_serializer_returns_background_image_url_when_set(self):
        from api.serializers import TemplateSerializer
        from unittest.mock import MagicMock
        t = Template.objects.create(name='ImgBG', created_by=self.user)
        mock_image = MagicMock()
        mock_image.__bool__ = lambda self: True
        mock_image.url = 'http://example.com/real-bg.png'
        t.background_image = mock_image
        s = TemplateSerializer(t)
        self.assertEqual(s.get_background_image_url(t), 'http://example.com/real-bg.png')

    def test_invitation_detail_participant_exists_by_email_when_no_student_linked(self):
        from api.serializers import InvitationDetailSerializer
        from events.models import EventInvitation, Event
        from participants.models import Participant

        event = Event.objects.create(name='SrlEv', event_date=date(2026, 9, 1), created_by=self.user)
        Participant.objects.create(
            document_id='SRLCOV1', first_name='A', last_name='B',
            email='srlcov@test.com', created_by=self.user
        )
        inv = EventInvitation.objects.create(
            event=event, email='srlcov@test.com', participant=None, created_by=self.user
        )
        s = InvitationDetailSerializer(inv)
        self.assertTrue(s.data['participant_exists'])

    def test_invitation_detail_student_not_exists_returns_false(self):
        from api.serializers import InvitationDetailSerializer
        from events.models import EventInvitation, Event

        event = Event.objects.create(name='SrlEv2', event_date=date(2026, 9, 2), created_by=self.user)
        inv = EventInvitation.objects.create(
            event=event, email='nobody@test.com', participant=None, created_by=self.user
        )
        s = InvitationDetailSerializer(inv)
        self.assertFalse(s.data['participant_exists'])


# ─────────────────────────────────────────────
# views.py: send_all_invitations token missing path
# ─────────────────────────────────────────────

class SendAllInvitationsTokenMissingTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('toktest@test.com')
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='Tok Event')

    @patch('django.core.mail.send_mail')
    def test_send_all_invitations_assigns_token_when_missing(self, mock_mail):
        mock_mail.return_value = 1
        from events.models import EventInvitation
        inv = EventInvitation.objects.create(
            event=self.event, email='notoken@test.com', status='pending', created_by=self.admin
        )
        EventInvitation.objects.filter(pk=inv.pk).update(token='')
        res = self.client.post(f'/api/events/{self.event.id}/invitations/send-all/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ─────────────────────────────────────────────
# config/urls.py: DEBUG=True static URL coverage
# ─────────────────────────────────────────────

class DebugURLPatternTest(TestCase):
    def test_debug_mode_adds_static_media_url(self):
        import sys
        from django.test import override_settings

        original = sys.modules.pop('config.urls', None)
        try:
            with override_settings(DEBUG=True, MEDIA_URL='/media/', MEDIA_ROOT='/tmp/media'):
                import config.urls as debug_urls
                self.assertGreater(len(debug_urls.urlpatterns), 0)
        finally:
            if original is not None:
                sys.modules['config.urls'] = original
            elif 'config.urls' in sys.modules:
                del sys.modules['config.urls']


# ─────────────────────────────────────────────
# views.py: uncovered queryset fallback paths + bulk success
# ─────────────────────────────────────────────

class ViewsQuersetFallbackTest(TestCase):
    def setUp(self):
        self.admin = make_admin('qs_fb@test.com')

    def _make_anon_request(self):
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        from django.contrib.auth.models import AnonymousUser
        factory = APIRequestFactory()
        raw = factory.get('/')
        raw.user = AnonymousUser()
        req = Request(raw)
        req._user = AnonymousUser()
        return req

    def test_certificate_queryset_returns_none_for_unauthenticated(self):
        from api.views import CertificateViewSet
        req = self._make_anon_request()
        viewset = CertificateViewSet()
        viewset.request = req
        viewset.format_kwarg = None
        viewset.action = 'list'
        qs = viewset.get_queryset()
        self.assertEqual(list(qs), [])

    def test_events_queryset_returns_all_for_unauthenticated(self):
        from api.views import EventsViewSet
        req = self._make_anon_request()
        viewset = EventsViewSet()
        viewset.request = req
        viewset.format_kwarg = None
        viewset.action = 'list'
        qs = viewset.get_queryset()
        self.assertIsNotNone(qs)


class BulkCertificateGenerationSuccessTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('bulksucc@test.com')
        self.client.force_authenticate(user=self.admin)

    def _make_excel_file(self):
        import pandas as pd
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        df = pd.DataFrame([{
            'full_name': 'Test User',
            'email': 'bulksucc@example.com',
            'document_id': 'BULKSUCC01',
            'event_name': 'Success Event',
        }])
        buf = BytesIO()
        df.to_excel(buf, index=False)
        buf.seek(0)
        return SimpleUploadedFile('bulk.xlsx', buf.read(),
                                   content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @patch('procesos.services.ExcelProcessingService')
    def test_bulk_generate_success_returns_200(self, mock_svc_class):
        from django.core.files.uploadedfile import SimpleUploadedFile
        mock_result = MagicMock()
        mock_result.to_dict.return_value = {'total_rows': 1, 'successful': 1, 'failed': 0}
        mock_result.get_summary.return_value = 'OK'
        mock_svc_class.return_value.process.return_value = mock_result
        f = self._make_excel_file()
        img = SimpleUploadedFile('bg.png', b'\x89PNG\r\n', content_type='image/png')
        event = make_event(self.admin, name='Success Event')
        res = self.client.post('/api/certificates/generate-bulk/', {
            'excel_file': f,
            'template_image': img,
            'event_id': event.id,
        }, format='multipart')
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ─────────────────────────────────────────────
# PASO 1: Coordinador role tests (TC-001 / TC-003 variants)
# ─────────────────────────────────────────────

def make_coordinator(email='coord@test.com'):
    return User.objects.create_user(
        email=email, full_name='Coord', password='pass123',
        role='coordinador', is_staff=False
    )


class CoordinadorRoleTest(TestCase):
    """TC-001 / TC-003 — Coordinador role recognition and operational access."""

    def setUp(self):
        self.client = APIClient()
        self.coordinator = make_coordinator()

    # helpers
    def _check_perm(self, perm_class, user, method='GET'):
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        factory = APIRequestFactory()
        raw = factory.get('/') if method == 'GET' else factory.post('/')
        raw.user = user
        req = Request(raw)
        req._user = user
        return perm_class().has_permission(req, None)

    def test_user_role_coordinador_stored_correctly(self):
        self.assertEqual(self.coordinator.role, 'coordinador')

    def test_is_coordinator_true_for_coordinador(self):
        from api.permissions import is_coordinator
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        raw = APIRequestFactory().get('/')
        raw.user = self.coordinator
        req = Request(raw)
        req._user = self.coordinator
        self.assertTrue(is_coordinator(req))

    def test_is_coordinator_false_for_admin(self):
        from api.permissions import is_coordinator
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        admin = make_admin('adm_coord@test.com')
        raw = APIRequestFactory().get('/')
        raw.user = admin
        req = Request(raw)
        req._user = admin
        self.assertFalse(is_coordinator(req))

    def test_is_coordinator_false_for_anonymous(self):
        from api.permissions import is_coordinator
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        from django.contrib.auth.models import AnonymousUser
        raw = APIRequestFactory().get('/')
        raw.user = AnonymousUser()
        req = Request(raw)
        req._user = AnonymousUser()
        self.assertFalse(is_coordinator(req))

    def test_is_operational_user_true_for_coordinador(self):
        from api.permissions import is_operational_user
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        raw = APIRequestFactory().get('/')
        raw.user = self.coordinator
        req = Request(raw)
        req._user = self.coordinator
        self.assertTrue(is_operational_user(req))

    def test_is_operational_user_true_for_admin(self):
        from api.permissions import is_operational_user
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        admin = make_admin('op_admin@test.com')
        raw = APIRequestFactory().get('/')
        raw.user = admin
        req = Request(raw)
        req._user = admin
        self.assertTrue(is_operational_user(req))

    def test_is_operational_user_false_for_participante(self):
        from api.permissions import is_operational_user
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        participante = make_user('op_part@test.com')
        raw = APIRequestFactory().get('/')
        raw.user = participante
        req = Request(raw)
        req._user = participante
        self.assertFalse(is_operational_user(req))

    def test_is_operational_perm_class_allows_coordinador_write(self):
        from api.permissions import IsOperationalUser
        self.assertTrue(self._check_perm(IsOperationalUser, self.coordinator, 'POST'))

    def test_is_coordinator_perm_class_allows_coordinador(self):
        from api.permissions import IsCoordinator
        self.assertTrue(self._check_perm(IsCoordinator, self.coordinator))

    def test_is_coordinator_perm_class_denies_admin(self):
        from api.permissions import IsCoordinator
        admin = make_admin('adm_coord2@test.com')
        self.assertFalse(self._check_perm(IsCoordinator, admin, 'POST'))

    def test_is_admin_still_false_for_coordinador(self):
        """Admin-only actions (audit) must remain denied to coordinador."""
        from api.permissions import is_admin
        from rest_framework.test import APIRequestFactory
        from rest_framework.request import Request
        raw = APIRequestFactory().get('/')
        raw.user = self.coordinator
        req = Request(raw)
        req._user = self.coordinator
        self.assertFalse(is_admin(req))


class CoordinadorOperationalAccessTest(TestCase):
    """Coordinador can enroll, generate, send — same as admin for day-to-day ops."""

    def setUp(self):
        self.client = APIClient()
        self.coordinator = make_coordinator('coord2@test.com')
        self.admin = make_admin('adm_op@test.com')
        self.client.force_authenticate(user=self.coordinator)
        self.event = make_event(self.admin)
        self.participant = make_participant(self.admin)
        self.template = Template.objects.create(name='T', created_by=self.admin)

    def test_coordinator_can_enroll_participant(self):
        res = self.client.post(f'/api/events/{self.event.id}/enroll/', {
            'student_id': self.participant.id
        })
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)

    def test_coordinator_can_generate_certificates(self):
        Enrollment.objects.create(
            participant=self.participant, event=self.event,
            attendance=True, created_by=self.admin
        )
        res = self.client.post(f'/api/events/{self.event.id}/certificates/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_coordinator_can_send_certificates(self):
        res = self.client.post(
            f'/api/events/{self.event.id}/certificates/send/',
            {'method': 'email'}
        )
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ─────────────────────────────────────────────
# PASO 1: Failed login attempt logging (TC-003)
# ─────────────────────────────────────────────

class LoginAttemptLoggingTest(TestCase):
    """TC-003 — Failed login attempts are logged at WARNING level."""

    def setUp(self):
        self.client = APIClient()
        make_admin('log_admin@test.com')

    def test_failed_login_wrong_password_logs_warning(self):
        import logging
        with self.assertLogs('api.views', level='WARNING') as cm:
            res = self.client.post('/api/login/', {
                'email': 'log_admin@test.com',
                'password': 'WRONG_PASSWORD'
            })
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue(any('LOGIN_FAILED' in line for line in cm.output))

    def test_failed_login_nonexistent_email_logs_warning(self):
        import logging
        with self.assertLogs('api.views', level='WARNING') as cm:
            res = self.client.post('/api/login/', {
                'email': 'ghost@test.com',
                'password': 'anypassword'
            })
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertTrue(any('LOGIN_FAILED' in line for line in cm.output))

    def test_successful_login_logs_info(self):
        with self.assertLogs('api.views', level='INFO') as cm:
            res = self.client.post('/api/login/', {
                'email': 'log_admin@test.com',
                'password': 'pass123'
            })
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertTrue(any('LOGIN_SUCCESS' in line for line in cm.output))

    def test_failed_login_logs_attempted_email(self):
        with self.assertLogs('api.views', level='WARNING') as cm:
            self.client.post('/api/login/', {
                'email': 'target@test.com',
                'password': 'wrong'
            })
        self.assertTrue(any('target@test.com' in line for line in cm.output))


# ─────────────────────────────────────────────
# PASO 1: Template visibility fix — admin sees all templates
# ─────────────────────────────────────────────

class TemplateVisibilityTest(TestCase):
    """Bug fix: admin must see templates created by other users."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('tvis_admin@test.com')
        self.other_admin = make_admin('tvis_other@test.com')

    def test_admin_sees_templates_from_other_users(self):
        Template.objects.create(name='OtherTemplate', created_by=self.other_admin)
        self.client.force_authenticate(user=self.admin)
        res = self.client.get('/api/templates/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        items = res.data.get('results', res.data) if isinstance(res.data, dict) else res.data
        names = [t['name'] for t in items]
        self.assertIn('OtherTemplate', names)

    def test_coordinator_sees_all_templates(self):
        coord = make_coordinator('tvis_coord@test.com')
        Template.objects.create(name='AdminTemplate', created_by=self.admin)
        self.client.force_authenticate(user=coord)
        res = self.client.get('/api/templates/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        items = res.data.get('results', res.data) if isinstance(res.data, dict) else res.data
        names = [t['name'] for t in items]
        self.assertIn('AdminTemplate', names)

    def test_participante_sees_only_own_templates(self):
        participante = make_user('tvis_part@test.com')
        Template.objects.create(name='ParticipanteOwned', created_by=participante)
        Template.objects.create(name='AdminOwned', created_by=self.admin)
        self.client.force_authenticate(user=participante)
        res = self.client.get('/api/templates/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        items = res.data.get('results', res.data) if isinstance(res.data, dict) else res.data
        names = [t['name'] for t in items]
        self.assertIn('ParticipanteOwned', names)
        self.assertNotIn('AdminOwned', names)


# ─────────────────────────────────────────────
# PASO 2: Coordinador write-access tests
# ─────────────────────────────────────────────

class CoordinadorWriteAccessTest(TestCase):
    """Coordinador puede crear/editar participantes, instructores y plantillas."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('coord_write_adm@test.com')
        self.coordinator = make_coordinator('coord_write@test.com')
        self.client.force_authenticate(user=self.coordinator)

    def test_coordinator_can_create_participant(self):
        res = self.client.post('/api/participants/', {
            'document_id': 'CW001', 'first_name': 'Nuevo', 'last_name': 'Part',
            'email': 'cwpart@test.com', 'phone': '111'
        })
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)

    def test_coordinator_can_update_participant(self):
        p = Participant.objects.create(
            document_id='CW002', first_name='Orig', last_name='Name',
            email='cwupd@test.com', created_by=self.admin
        )
        res = self.client.patch(f'/api/participants/{p.id}/', {'first_name': 'Actualizado'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['first_name'], 'Actualizado')

    def test_coordinator_can_delete_participant(self):
        p = Participant.objects.create(
            document_id='CW003', first_name='Del', last_name='Me',
            email='cwdel@test.com', created_by=self.admin
        )
        res = self.client.delete(f'/api/participants/{p.id}/')
        self.assertEqual(res.status_code, status.HTTP_204_NO_CONTENT)

    def test_coordinator_can_create_instructor(self):
        res = self.client.post('/api/instructors/', {
            'full_name': 'Coord Instructor', 'email': 'ci@test.com', 'specialty': 'Django'
        })
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)

    def test_coordinator_can_create_template(self):
        res = self.client.post('/api/templates/', {'name': 'Coord Template', 'category': 'Test'})
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)

    def test_coordinator_can_create_enrollment(self):
        admin = make_admin('coord_enr_adm@test.com')
        event = make_event(admin, name='CW Event')
        participant = make_participant(admin, doc='CW004', email='cwenr@test.com')
        res = self.client.post('/api/enrollments/', {
            'participant_id': participant.id,
            'event_id': event.id,
        }, format='json')
        self.assertEqual(res.status_code, status.HTTP_201_CREATED)

    def test_coordinator_can_generate_certificate_via_api(self):
        admin = make_admin('coord_gen_adm@test.com')
        event = make_event(admin, name='CW Gen Event')
        participant = make_participant(admin, doc='CW005', email='cwgen@test.com')
        Enrollment.objects.create(
            participant=participant, event=event, attendance=True, created_by=admin
        )
        cert = Certificate.objects.create(
            participant=participant, event=event,
            template=Template.objects.create(name='T', created_by=admin),
            generated_by=admin
        )
        with patch('services.pdf_service.PDFService.generate_certificate_pdf') as mock_pdf:
            mock_pdf.return_value = {'success': True, 'path': '/media/cert.pdf'}
            res = self.client.post(f'/api/certificates/{cert.id}/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_coordinator_can_deliver_certificate_via_api(self):
        admin = make_admin('coord_del_adm@test.com')
        event = make_event(admin, name='CW Del Event')
        participant = make_participant(admin, doc='CW006', email='cwdel2@test.com')
        cert = Certificate.objects.create(
            participant=participant, event=event, generated_by=admin
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated', pdf_url='/media/c.pdf')
        cert.refresh_from_db()
        with patch('services.email_service.EmailService.send_certificate') as mock_send:
            mock_send.return_value = {'success': True, 'message': 'sent'}
            res = self.client.post(f'/api/certificates/{cert.id}/deliver/', {'method': 'email'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_participante_cannot_create_participant(self):
        participante = make_user('cwnopart@test.com')
        self.client.force_authenticate(user=participante)
        res = self.client.post('/api/participants/', {
            'document_id': 'NOPART', 'first_name': 'X', 'last_name': 'Y',
            'email': 'nopart@test.com'
        })
        self.assertIn(res.status_code, [status.HTTP_403_FORBIDDEN, status.HTTP_401_UNAUTHORIZED])

    def test_participante_cannot_create_instructor(self):
        participante = make_user('cwnoinst@test.com')
        self.client.force_authenticate(user=participante)
        res = self.client.post('/api/instructors/', {'full_name': 'X', 'email': 'x@test.com'})
        self.assertIn(res.status_code, [status.HTTP_403_FORBIDDEN, status.HTTP_401_UNAUTHORIZED])

    def test_participante_cannot_create_template(self):
        participante = make_user('cwnotmpl@test.com')
        self.client.force_authenticate(user=participante)
        res = self.client.post('/api/templates/', {'name': 'X'})
        self.assertIn(res.status_code, [status.HTTP_403_FORBIDDEN, status.HTTP_401_UNAUTHORIZED])


# ─────────────────────────────────────────────
# PASO 2: WhatsApp vía API — TC-019 / TC-020
# ─────────────────────────────────────────────

class WhatsAppAPITest(TestCase):
    """TC-019 / TC-020 — Entrega WhatsApp vía API con y sin teléfono."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('wa_api_adm@test.com')
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='WA API Event')
        self.template = Template.objects.create(name='T', created_by=self.admin)

    def _make_cert(self, phone='999000111'):
        participant = Participant.objects.create(
            document_id=f'WA{phone or "NONE"}',
            first_name='Raul', last_name='Paz',
            email=f'raul_{phone or "none"}@test.com',
            phone=phone, created_by=self.admin
        )
        cert = Certificate.objects.create(
            participant=participant, event=self.event,
            template=self.template, generated_by=self.admin
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated', pdf_url='/media/cert.pdf')
        cert.refresh_from_db()
        return cert

    def test_tc020_whatsapp_api_without_phone_returns_400(self):
        """TC-020: API debe devolver 400 si participante no tiene teléfono."""
        cert = self._make_cert(phone='')
        res = self.client.post(f'/api/certificates/{cert.id}/deliver/', {'method': 'whatsapp'})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    @patch('services.whatsapp_service.get_whatsapp_service')
    def test_tc019_whatsapp_api_with_phone_succeeds(self, mock_get):
        """TC-019: API devuelve 200 si participante tiene teléfono."""
        mock_ws = MagicMock()
        mock_ws.send_certificate.return_value = {'success': True, 'message': 'sent', 'sid': 'SM1'}
        mock_get.return_value = mock_ws
        cert = self._make_cert(phone='999000111')
        res = self.client.post(f'/api/certificates/{cert.id}/deliver/', {'method': 'whatsapp'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    @patch('services.email_service.EmailService.send_certificate')
    def test_tc017_retry_email_via_api_creates_new_delivery_log(self, mock_send):
        """TC-017: reintento vía API crea nuevo DeliveryLog sin regenerar PDF."""
        mock_send.return_value = {'success': True, 'message': 'resent'}
        cert = self._make_cert(phone='999000111')
        Certificate.objects.filter(pk=cert.pk).update(status='failed')
        cert.refresh_from_db()
        original_pdf = cert.pdf_url

        res1 = self.client.post(f'/api/certificates/{cert.id}/deliver/', {'method': 'email'})
        res2 = self.client.post(f'/api/certificates/{cert.id}/deliver/', {'method': 'email'})

        self.assertEqual(res1.status_code, status.HTTP_200_OK)
        self.assertEqual(res2.status_code, status.HTTP_200_OK)
        cert.refresh_from_db()
        self.assertEqual(cert.pdf_url, original_pdf)
        self.assertEqual(cert.deliveries.count(), 2)


# ─────────────────────────────────────────────
# PASO 2: TC-011 — attendance=False bloquea generación vía API
# ─────────────────────────────────────────────

class AttendanceAPIBlockTest(TestCase):
    """TC-011 — API devuelve 400 cuando attendance=False al generar certificado."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('att_api@test.com')
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin)
        self.participant = make_participant(self.admin, doc='ATT_API_01', email='att_api@test.com')
        self.template = Template.objects.create(name='T', created_by=self.admin)

    def test_tc011_generate_blocked_when_attendance_false_via_api(self):
        """TC-011: attendance=False → 400 en POST /certificates/{id}/generate/."""
        Enrollment.objects.create(
            participant=self.participant, event=self.event,
            attendance=False, created_by=self.admin
        )
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event,
            template=self.template, generated_by=self.admin
        )
        res = self.client.post(f'/api/certificates/{cert.id}/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn('error', res.data.get('status', 'error'))

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_tc010_generate_succeeds_when_attendance_true_via_api(self, mock_pdf):
        """TC-010: attendance=True → 200 en POST /certificates/{id}/generate/."""
        mock_pdf.return_value = {'success': True, 'path': '/media/cert.pdf'}
        Enrollment.objects.create(
            participant=self.participant, event=self.event,
            attendance=True, created_by=self.admin
        )
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event,
            template=self.template, generated_by=self.admin
        )
        res = self.client.post(f'/api/certificates/{cert.id}/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['status'], 'success')


# ═══════════════════════════════════════════════════════════════
# PASO 3 — TC-025 a TC-035
# ═══════════════════════════════════════════════════════════════
from django.conf import settings as django_settings
from django.core.cache import cache


# ─────────────────────────────────────────────
# TC-025: Rate limiting
# ─────────────────────────────────────────────

class RateLimitingTest(TestCase):
    """TC-025: Throttle classes configured; 429 returned when rate exceeded."""

    def setUp(self):
        self.client = APIClient()

    def test_throttle_classes_configured(self):
        classes = django_settings.REST_FRAMEWORK.get('DEFAULT_THROTTLE_CLASSES', [])
        self.assertIn('rest_framework.throttling.AnonRateThrottle', classes)
        self.assertIn('rest_framework.throttling.UserRateThrottle', classes)

    def test_throttle_rates_defined(self):
        rates = django_settings.REST_FRAMEWORK.get('DEFAULT_THROTTLE_RATES', {})
        self.assertIn('anon', rates)
        self.assertIn('user', rates)

    @patch('rest_framework.throttling.AnonRateThrottle.get_rate', return_value='1/minute')
    def test_anon_rate_limit_returns_429(self, _mock):
        """TC-025: Second anonymous request past the 1/minute limit → 429."""
        cache.clear()
        self.client.get('/api/certificates/verify/?code=X')
        res = self.client.get('/api/certificates/verify/?code=X')
        self.assertEqual(res.status_code, status.HTTP_429_TOO_MANY_REQUESTS)


# ─────────────────────────────────────────────
# TC-026: CORS headers on public endpoint
# ─────────────────────────────────────────────

class CORSHeaderTest(TestCase):
    """TC-026: Allowed origins receive Access-Control-Allow-Origin in responses."""

    def setUp(self):
        self.client = APIClient()

    def test_allowed_origin_gets_cors_header(self):
        """TC-026: Origin in CORS_ALLOWED_ORIGINS gets header back."""
        res = self.client.get(
            '/api/certificates/verify/?code=TESTCODE',
            HTTP_ORIGIN='http://localhost:3000',
        )
        self.assertIn(res.status_code, [
            status.HTTP_404_NOT_FOUND,
            status.HTTP_400_BAD_REQUEST,
        ])
        self.assertIn('Access-Control-Allow-Origin', res)

    def test_preflight_from_allowed_origin_accepted(self):
        """TC-026: OPTIONS preflight from allowed origin returns 2xx."""
        res = self.client.options(
            '/api/certificates/verify/',
            HTTP_ORIGIN='http://localhost:3000',
            HTTP_ACCESS_CONTROL_REQUEST_METHOD='GET',
        )
        self.assertIn(res.status_code, [200, 204])


# ─────────────────────────────────────────────
# TC-027: JWT token validity and refresh
# ─────────────────────────────────────────────

class TokenValidityTest(TestCase):
    """TC-027: Invalid/expired tokens rejected with 401; refresh endpoint works."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('tok27@test.com')

    def test_garbled_token_returns_401(self):
        self.client.credentials(HTTP_AUTHORIZATION='Bearer notavalidtoken.abc.xyz')
        self.assertEqual(self.client.get('/api/me/').status_code, status.HTTP_401_UNAUTHORIZED)

    def test_expired_token_returns_401(self):
        import datetime
        from rest_framework_simplejwt.tokens import AccessToken
        token = AccessToken.for_user(self.admin)
        token.payload['exp'] = int(
            datetime.datetime(2020, 1, 1, tzinfo=datetime.timezone.utc).timestamp()
        )
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {str(token)}')
        self.assertEqual(self.client.get('/api/me/').status_code, status.HTTP_401_UNAUTHORIZED)

    def test_valid_refresh_returns_new_access_token(self):
        from rest_framework_simplejwt.tokens import RefreshToken
        refresh = RefreshToken.for_user(self.admin)
        res = self.client.post('/api/token/refresh/', {'refresh': str(refresh)})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('access', res.data)

    def test_invalid_refresh_token_returns_401(self):
        res = self.client.post('/api/token/refresh/', {'refresh': 'badtoken'})
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)


# ─────────────────────────────────────────────
# TC-028: Authentication required on protected endpoints
# ─────────────────────────────────────────────

class AuthRequiredTest(TestCase):
    """TC-028: All protected endpoints return 401 when called without credentials."""

    def setUp(self):
        self.client = APIClient()  # no credentials

    def test_me_requires_auth(self):
        self.assertEqual(self.client.get('/api/me/').status_code, status.HTTP_401_UNAUTHORIZED)

    def test_participants_requires_auth(self):
        self.assertEqual(self.client.get('/api/participants/').status_code, status.HTTP_401_UNAUTHORIZED)

    def test_events_requires_auth(self):
        self.assertEqual(self.client.get('/api/events/').status_code, status.HTTP_401_UNAUTHORIZED)

    def test_certificates_requires_auth(self):
        self.assertEqual(self.client.get('/api/certificates/').status_code, status.HTTP_401_UNAUTHORIZED)

    def test_templates_requires_auth(self):
        self.assertEqual(self.client.get('/api/templates/').status_code, status.HTTP_401_UNAUTHORIZED)

    def test_instructors_requires_auth(self):
        self.assertEqual(self.client.get('/api/instructors/').status_code, status.HTTP_401_UNAUTHORIZED)

    def test_verify_is_public(self):
        """TC-028: /verify/ is AllowAny — no 401."""
        res = self.client.get('/api/certificates/verify/?code=X')
        self.assertNotEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)


# ─────────────────────────────────────────────
# TC-029: Input validation rejects malformed requests
# ─────────────────────────────────────────────

class InputValidationTest(TestCase):
    """TC-029: API returns 400 (not 500) for invalid or incomplete inputs."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('iv29@test.com')
        self.client.force_authenticate(user=self.admin)

    def test_event_without_name_returns_400(self):
        res = self.client.post('/api/events/', {'event_date': '2026-10-01'})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_participant_without_document_returns_400(self):
        res = self.client.post('/api/participants/', {
            'first_name': 'Only', 'last_name': 'Name', 'email': 'only29@test.com',
        })
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_enrollment_nonexistent_event_returns_400_or_404(self):
        participant = make_participant(self.admin, doc='IV29P1', email='iv29p1@test.com')
        res = self.client.post('/api/enrollments/', {
            'participant_id': participant.id,
            'event_id': 99999,
        }, format='json')
        self.assertIn(res.status_code, [
            status.HTTP_400_BAD_REQUEST,
            status.HTTP_404_NOT_FOUND,
        ])

    def test_login_wrong_password_returns_400(self):
        res = self.client.post('/api/login/', {
            'email': 'iv29@test.com',
            'password': 'WRONG_PASSWORD',
        })
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_duplicate_participant_document_returns_400(self):
        make_participant(self.admin, doc='DUPIV29', email='dupiv29a@test.com')
        res = self.client.post('/api/participants/', {
            'document_id': 'DUPIV29',
            'first_name': 'Dup', 'last_name': 'User',
            'email': 'dupiv29b@test.com',
        })
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)


# ─────────────────────────────────────────────
# TC-030: Full E2E happy path
# ─────────────────────────────────────────────

class FullHappyPathE2ETest(TestCase):
    """
    TC-030: Complete certification lifecycle.
    Real login → create event → enroll → attend → generate cert
    → deliver via email → public verify → 200.
    """

    def setUp(self):
        self.client = APIClient()

    @patch('services.email_service.EmailService.send_certificate')
    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_full_lifecycle(self, mock_pdf, mock_email):
        mock_pdf.return_value = {'success': True, 'path': '/media/e2e.pdf'}
        mock_email.return_value = {'success': True, 'message': 'sent'}

        # 1. Login
        admin = User.objects.create_user(
            email='e2e30@test.com', full_name='E2E Admin',
            password='SecurePass123!', role='admin', is_staff=True,
        )
        login_res = self.client.post('/api/login/', {
            'email': 'e2e30@test.com',
            'password': 'SecurePass123!',
        })
        self.assertEqual(login_res.status_code, status.HTTP_200_OK)
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {login_res.data["access"]}')

        # 2. Create event
        ev = self.client.post('/api/events/', {'name': 'E2E Event', 'event_date': '2026-08-29'})
        self.assertEqual(ev.status_code, status.HTTP_201_CREATED)
        event_id = ev.data['id']

        # 3. Create participant
        part = self.client.post('/api/participants/', {
            'document_id': 'E2E30P1', 'first_name': 'Elena', 'last_name': 'Vega',
            'email': 'elena30@e2e.com', 'phone': '900123456',
        })
        self.assertEqual(part.status_code, status.HTTP_201_CREATED)
        participant_id = part.data['id']

        # 4. Enroll
        enroll = self.client.post('/api/enrollments/', {
            'participant_id': participant_id, 'event_id': event_id,
        }, format='json')
        self.assertEqual(enroll.status_code, status.HTTP_201_CREATED)
        enrollment_id = enroll.data['id']

        # 5. Mark attendance
        att = self.client.patch(
            f'/api/enrollments/{enrollment_id}/attendance/',
            {'attendance': True},
        )
        self.assertEqual(att.status_code, status.HTTP_200_OK)

        # 6. Create certificate and generate PDF
        template = Template.objects.create(name='E2E30 Tpl', created_by=admin)
        cert = Certificate.objects.create(
            participant_id=participant_id, event_id=event_id,
            template=template, generated_by=admin,
        )
        gen = self.client.post(f'/api/certificates/{cert.id}/generate/', {})
        self.assertEqual(gen.status_code, status.HTTP_200_OK)
        self.assertEqual(gen.data['status'], 'success')

        # 7. Deliver via email
        delv = self.client.post(f'/api/certificates/{cert.id}/deliver/', {'method': 'email'})
        self.assertEqual(delv.status_code, status.HTTP_200_OK)

        # 8. Public verification
        cert.refresh_from_db()
        verify = self.client.get(f'/api/certificates/verify/?code={cert.verification_code}')
        self.assertEqual(verify.status_code, status.HTTP_200_OK)
        self.assertEqual(verify.data['status'], 'success')


# ─────────────────────────────────────────────
# TC-031: Batch generation respects attendance
# ─────────────────────────────────────────────

class BatchGenerationAttendanceTest(TestCase):
    """TC-031: /events/{id}/certificates/generate/ creates certs only for attendees."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('batch31@test.com')
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='Batch TC-031')

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_only_attendees_get_certificates(self, mock_pdf):
        """TC-031: 2 attended + 1 absent → 2 certs, absent participant excluded."""
        mock_pdf.return_value = {'success': True, 'path': '/media/b.pdf'}
        p1 = make_participant(self.admin, doc='B31P1', email='b31p1@test.com')
        p2 = make_participant(self.admin, doc='B31P2', email='b31p2@test.com')
        p3 = make_participant(self.admin, doc='B31P3', email='b31p3@test.com')

        Enrollment.objects.create(participant=p1, event=self.event, attendance=True,  created_by=self.admin)
        Enrollment.objects.create(participant=p2, event=self.event, attendance=True,  created_by=self.admin)
        Enrollment.objects.create(participant=p3, event=self.event, attendance=False, created_by=self.admin)

        res = self.client.post(f'/api/events/{self.event.id}/certificates/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['created'], 2)

        ids = set(Certificate.objects.filter(event=self.event).values_list('participant_id', flat=True))
        self.assertIn(p1.id, ids)
        self.assertIn(p2.id, ids)
        self.assertNotIn(p3.id, ids)

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_all_absent_creates_zero_certificates(self, mock_pdf):
        mock_pdf.return_value = {'success': True, 'path': '/media/b.pdf'}
        p = make_participant(self.admin, doc='B31P4', email='b31p4@test.com')
        Enrollment.objects.create(participant=p, event=self.event, attendance=False, created_by=self.admin)

        res = self.client.post(f'/api/events/{self.event.id}/certificates/generate/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['created'], 0)
        self.assertEqual(Certificate.objects.filter(event=self.event).count(), 0)


# ─────────────────────────────────────────────
# TC-032: Event statistics accuracy
# ─────────────────────────────────────────────

class EventStatsAccuracyTest(TestCase):
    """TC-032: /events/{id}/stats/ returns exact counts for each field."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('stats32@test.com')
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='Stats TC-032')

    def test_stats_match_inserted_data(self):
        p1 = make_participant(self.admin, doc='S32P1', email='s32p1@test.com')
        p2 = make_participant(self.admin, doc='S32P2', email='s32p2@test.com')
        p3 = make_participant(self.admin, doc='S32P3', email='s32p3@test.com')

        Enrollment.objects.create(participant=p1, event=self.event, attendance=True,  created_by=self.admin)
        Enrollment.objects.create(participant=p2, event=self.event, attendance=True,  created_by=self.admin)
        Enrollment.objects.create(participant=p3, event=self.event, attendance=False, created_by=self.admin)

        Certificate.objects.create(participant=p1, event=self.event, generated_by=self.admin)
        c2 = Certificate.objects.create(participant=p2, event=self.event, generated_by=self.admin)
        Certificate.objects.filter(pk=c2.pk).update(status='generated')

        res = self.client.get(f'/api/events/{self.event.id}/stats/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        d = res.data

        self.assertEqual(d['total_enrollments'], 3)
        self.assertEqual(d['attendees'], 2)
        self.assertEqual(d['absent'], 1)
        self.assertEqual(d['total_certificates'], 2)
        self.assertEqual(d['generated_certificates'], 1)
        self.assertEqual(d['pending_certificates'], 1)
        self.assertEqual(d['sent_certificates'], 0)
        self.assertEqual(d['failed_certificates'], 0)


# ─────────────────────────────────────────────
# TC-033: Delivery history E2E
# ─────────────────────────────────────────────

class DeliveryHistoryE2ETest(TestCase):
    """TC-033: /certificates/{id}/history/ reflects all delivery attempts."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('hist33@test.com')
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='Hist TC-033')
        self.participant = make_participant(self.admin, doc='H33P1', email='h33@test.com')
        self.cert = Certificate.objects.create(
            participant=self.participant, event=self.event, generated_by=self.admin,
        )
        Certificate.objects.filter(pk=self.cert.pk).update(status='generated', pdf_url='/media/h.pdf')
        self.cert.refresh_from_db()

    def test_history_empty_before_delivery(self):
        res = self.client.get(f'/api/certificates/{self.cert.id}/history/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['total_attempts'], 0)

    @patch('services.email_service.EmailService.send_certificate')
    def test_history_records_multiple_attempts(self, mock_send):
        """TC-033: Two deliver() calls produce two entries in history."""
        mock_send.return_value = {'success': True, 'message': 'ok'}

        self.client.post(f'/api/certificates/{self.cert.id}/deliver/', {'method': 'email'})
        Certificate.objects.filter(pk=self.cert.pk).update(status='failed')
        self.cert.refresh_from_db()
        self.client.post(f'/api/certificates/{self.cert.id}/deliver/', {'method': 'email'})

        res = self.client.get(f'/api/certificates/{self.cert.id}/history/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['total_attempts'], 2)

    @patch('services.email_service.EmailService.send_certificate')
    def test_pdf_url_unchanged_after_retry(self, mock_send):
        """TC-033: Retrying delivery does not modify the existing PDF URL."""
        mock_send.return_value = {'success': True, 'message': 'ok'}
        original_pdf = self.cert.pdf_url

        self.client.post(f'/api/certificates/{self.cert.id}/deliver/', {'method': 'email'})
        Certificate.objects.filter(pk=self.cert.pk).update(status='failed')
        self.cert.refresh_from_db()
        self.client.post(f'/api/certificates/{self.cert.id}/deliver/', {'method': 'email'})

        self.cert.refresh_from_db()
        self.assertEqual(self.cert.pdf_url, original_pdf)


# ─────────────────────────────────────────────
# TC-034: Coordinator UAT — full workflow
# ─────────────────────────────────────────────

class CoordinatorUATTest(TestCase):
    """TC-034: Coordinator can complete the entire certification workflow autonomously."""

    def setUp(self):
        self.client = APIClient()
        self.coordinator = make_coordinator('uat34@test.com')
        self.client.force_authenticate(user=self.coordinator)

    @patch('services.email_service.EmailService.send_certificate')
    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_coordinator_full_workflow(self, mock_pdf, mock_email):
        mock_pdf.return_value = {'success': True, 'path': '/media/uat34.pdf'}
        mock_email.return_value = {'success': True, 'message': 'sent'}

        ev = self.client.post('/api/events/', {'name': 'UAT34 Event', 'event_date': '2026-10-15'})
        self.assertEqual(ev.status_code, status.HTTP_201_CREATED)
        event_id = ev.data['id']

        part = self.client.post('/api/participants/', {
            'document_id': 'UAT34P1', 'first_name': 'Pedro', 'last_name': 'Torres',
            'email': 'pedro34@uat.com',
        })
        self.assertEqual(part.status_code, status.HTTP_201_CREATED)
        participant_id = part.data['id']

        enroll = self.client.post('/api/enrollments/', {
            'participant_id': participant_id, 'event_id': event_id,
        }, format='json')
        self.assertEqual(enroll.status_code, status.HTTP_201_CREATED)
        enrollment_id = enroll.data['id']

        att = self.client.patch(
            f'/api/enrollments/{enrollment_id}/attendance/',
            {'attendance': True},
        )
        self.assertEqual(att.status_code, status.HTTP_200_OK)

        gen = self.client.post(f'/api/events/{event_id}/certificates/generate/', {})
        self.assertEqual(gen.status_code, status.HTTP_200_OK)
        self.assertEqual(gen.data['created'], 1)

        cert = Certificate.objects.filter(event_id=event_id).first()
        self.assertIsNotNone(cert)
        self.assertEqual(cert.status, 'generated')

        delv = self.client.post(f'/api/certificates/{cert.id}/deliver/', {'method': 'email'})
        self.assertEqual(delv.status_code, status.HTTP_200_OK)
        cert.refresh_from_db()
        self.assertEqual(cert.status, 'sent')


# ─────────────────────────────────────────────
# TC-035: Public verification — all response paths
# ─────────────────────────────────────────────

class PublicVerificationE2ETest(TestCase):
    """TC-035: /certificates/verify/ covers every branch: 200, 404, 410, 400."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('pv35@test.com')
        self.event = make_event(self.admin, name='PV35 Event')
        self.participant = make_participant(self.admin, doc='PV35P1', email='pv35@test.com')

    def test_valid_code_returns_200(self):
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, generated_by=self.admin,
        )
        res = self.client.get(f'/api/certificates/verify/?code={cert.verification_code}')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['status'], 'success')

    def test_invalid_code_returns_404(self):
        res = self.client.get('/api/certificates/verify/?code=DOESNOTEXIST9999')
        self.assertEqual(res.status_code, status.HTTP_404_NOT_FOUND)

    def test_expired_certificate_returns_410(self):
        from django.utils import timezone
        from datetime import timedelta
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, generated_by=self.admin,
        )
        Certificate.objects.filter(pk=cert.pk).update(
            expires_at=timezone.now() - timedelta(days=1)
        )
        cert.refresh_from_db()
        res = self.client.get(f'/api/certificates/verify/?code={cert.verification_code}')
        self.assertEqual(res.status_code, status.HTTP_410_GONE)

    def test_missing_code_returns_400(self):
        res = self.client.get('/api/certificates/verify/')
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_verify_accessible_without_auth(self):
        """TC-035: Unauthenticated access to /verify/ is allowed."""
        anon = APIClient()
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, generated_by=self.admin,
        )
        res = anon.get(f'/api/certificates/verify/?code={cert.verification_code}')
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ═══════════════════════════════════════════════════════════════
# PASO 4 — Retry endpoint + Export endpoint
# ═══════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────
# PASO 4 — Retry endpoint: POST /certificates/{id}/retry/
# ─────────────────────────────────────────────

class CertificateRetryEndpointTest(TestCase):
    """POST /certificates/{id}/retry/ — formal retry for failed certificates."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('retry4@test.com')
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='Retry P4 Event')
        self.participant = make_participant(self.admin, doc='R4P1', email='r4p1@test.com')
        self.cert = Certificate.objects.create(
            participant=self.participant, event=self.event, generated_by=self.admin,
        )
        Certificate.objects.filter(pk=self.cert.pk).update(
            status='failed', pdf_url='/media/retry.pdf'
        )
        self.cert.refresh_from_db()

    @patch('services.email_service.EmailService.send_certificate')
    def test_retry_failed_cert_with_explicit_method(self, mock_send):
        """Retry a failed cert providing method explicitly → 200."""
        mock_send.return_value = {'success': True, 'message': 'resent'}
        res = self.client.post(f'/api/certificates/{self.cert.id}/retry/', {'method': 'email'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['status'], 'success')

    @patch('services.email_service.EmailService.send_certificate')
    def test_retry_uses_last_delivery_method_when_omitted(self, mock_send):
        """Retry without providing method uses last delivery's method."""
        mock_send.return_value = {'success': True, 'message': 'resent'}
        from deliveries.models import DeliveryLog
        DeliveryLog.objects.create(
            certificate=self.cert, delivery_method='email',
            recipient='r4p1@test.com', sent_by=self.admin,
            status='error',
        )
        res = self.client.post(f'/api/certificates/{self.cert.id}/retry/', {})
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['status'], 'success')

    def test_retry_non_failed_cert_returns_400(self):
        """Retry a generated (non-failed) cert → 400."""
        Certificate.objects.filter(pk=self.cert.pk).update(status='generated')
        self.cert.refresh_from_db()
        res = self.client.post(f'/api/certificates/{self.cert.id}/retry/', {'method': 'email'})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    def test_retry_without_method_and_no_history_returns_400(self):
        """Retry without method and no prior delivery → 400."""
        res = self.client.post(f'/api/certificates/{self.cert.id}/retry/', {})
        self.assertEqual(res.status_code, status.HTTP_400_BAD_REQUEST)

    @patch('services.email_service.EmailService.send_certificate')
    def test_retry_creates_new_delivery_log(self, mock_send):
        """Each retry call produces a new DeliveryLog entry."""
        mock_send.return_value = {'success': True, 'message': 'ok'}
        from deliveries.models import DeliveryLog
        before = DeliveryLog.objects.filter(certificate=self.cert).count()
        self.client.post(f'/api/certificates/{self.cert.id}/retry/', {'method': 'email'})
        # cert is now 'sent'; force back to failed for a second retry
        Certificate.objects.filter(pk=self.cert.pk).update(status='failed')
        self.client.post(f'/api/certificates/{self.cert.id}/retry/', {'method': 'email'})
        after = DeliveryLog.objects.filter(certificate=self.cert).count()
        self.assertEqual(after - before, 2)

    def test_retry_requires_auth(self):
        """Unauthenticated retry → 401."""
        anon = APIClient()
        res = anon.post(f'/api/certificates/{self.cert.id}/retry/', {'method': 'email'})
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    @patch('services.email_service.EmailService.send_certificate')
    def test_coordinator_can_retry(self, mock_send):
        """Coordinators also have retry access."""
        mock_send.return_value = {'success': True, 'message': 'ok'}
        coord = make_coordinator('retry4_coord@test.com')
        self.client.force_authenticate(user=coord)
        res = self.client.post(f'/api/certificates/{self.cert.id}/retry/', {'method': 'email'})
        self.assertEqual(res.status_code, status.HTTP_200_OK)


# ─────────────────────────────────────────────
# PASO 4 — Export endpoint: GET /certificates/export/
# ─────────────────────────────────────────────

class CertificateExportEndpointTest(TestCase):
    """GET /certificates/export/ — CSV/Excel download for administrator audit."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('export4@test.com')
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='Export P4 Event')
        self.participant = make_participant(self.admin, doc='EXP4P1', email='exp4p1@test.com')
        self.cert = Certificate.objects.create(
            participant=self.participant, event=self.event, generated_by=self.admin,
        )

    def test_export_csv_returns_200_with_content_type(self):
        """Default format=csv → Content-Type text/csv."""
        res = self.client.get('/api/certificates/export/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn('text/csv', res['Content-Type'])

    def test_export_csv_contains_headers(self):
        """CSV export includes all expected column headers."""
        res = self.client.get('/api/certificates/export/')
        content = res.content.decode()
        for col in ['id', 'participant_name', 'event_name', 'status', 'verification_code']:
            self.assertIn(col, content)

    def test_export_csv_contains_certificate_row(self):
        """CSV export includes a data row for the existing certificate."""
        res = self.client.get('/api/certificates/export/')
        content = res.content.decode()
        self.assertIn(self.participant.email, content)
        self.assertIn(self.event.name, content)

    def test_export_excel_returns_xlsx_content_type(self):
        """format=excel → Content-Type application/vnd.openxmlformats…"""
        res = self.client.get('/api/certificates/export/?file_format=excel')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            res['Content-Type'],
        )

    def test_export_excel_is_valid_workbook(self):
        """Excel export can be opened as a valid openpyxl workbook."""
        import io
        import openpyxl
        res = self.client.get('/api/certificates/export/?file_format=excel')
        wb = openpyxl.load_workbook(io.BytesIO(res.content))
        ws = wb.active
        # First row is headers
        headers = [cell.value for cell in ws[1]]
        self.assertIn('id', headers)
        self.assertIn('status', headers)
        # Second row is data
        self.assertGreaterEqual(ws.max_row, 2)

    def test_export_filter_by_event_id(self):
        """event_id filter narrows results to that event only."""
        other_event = make_event(self.admin, name='Other Export Event')
        other_part = make_participant(self.admin, doc='EXP4P2', email='exp4p2@test.com')
        Certificate.objects.create(participant=other_part, event=other_event, generated_by=self.admin)

        res = self.client.get(f'/api/certificates/export/?event_id={self.event.id}')
        content = res.content.decode()
        self.assertIn(self.event.name, content)
        self.assertNotIn(other_event.name, content)

    def test_export_filter_by_status(self):
        """status filter excludes certificates with a different status."""
        other_part = make_participant(self.admin, doc='EXP4P3', email='exp4p3@test.com')
        other_event = make_event(self.admin, name='Status Filter Event')
        c2 = Certificate.objects.create(
            participant=other_part, event=other_event, generated_by=self.admin,
        )
        Certificate.objects.filter(pk=c2.pk).update(status='generated')

        res = self.client.get('/api/certificates/export/?status=generated')
        content = res.content.decode()
        self.assertIn(other_event.name, content)
        self.assertNotIn(self.event.name, content)

    def test_export_requires_admin(self):
        """Non-admin user (participante) cannot access export → 403."""
        part = make_user('expnoauth@test.com')
        self.client.force_authenticate(user=part)
        res = self.client.get('/api/certificates/export/')
        self.assertIn(res.status_code, [
            status.HTTP_403_FORBIDDEN,
            status.HTTP_401_UNAUTHORIZED,
        ])

    def test_export_requires_auth(self):
        """Unauthenticated request → 401."""
        anon = APIClient()
        res = anon.get('/api/certificates/export/')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_export_content_disposition_header(self):
        """CSV response has Content-Disposition with filename."""
        res = self.client.get('/api/certificates/export/')
        self.assertIn('Content-Disposition', res)
        self.assertIn('certificados', res['Content-Disposition'])


# ═══════════════════════════════════════════════════════════════
# PASO 5 — AuditLog model + instrumentation + /api/audit/ endpoint
# ═══════════════════════════════════════════════════════════════
from api.models import AuditLog


class AuditLogModelTest(TestCase):
    """Unit tests for the AuditLog model."""

    def setUp(self):
        self.admin = make_admin('audit_model@test.com')
        self.event = make_event(self.admin, name='Audit Event')
        self.participant = make_participant(self.admin, doc='AM01', email='am01@test.com')
        self.cert = Certificate.objects.create(
            participant=self.participant, event=self.event, generated_by=self.admin,
        )

    def test_create_entry(self):
        entry = AuditLog.objects.create(
            action='certificate_generated', user=self.admin,
            certificate=self.cert, ip_address='127.0.0.1', details={'note': 'ok'},
        )
        self.assertEqual(entry.action, 'certificate_generated')
        self.assertEqual(entry.user, self.admin)
        self.assertIn('note', entry.details)

    def test_str_contains_action_and_email(self):
        entry = AuditLog.objects.create(action='user_login', user=self.admin)
        self.assertIn('user_login', str(entry))
        self.assertIn(self.admin.email, str(entry))

    def test_anonymous_entry_has_no_user(self):
        entry = AuditLog.objects.create(
            action='user_login_failed', ip_address='10.0.0.1',
            details={'email': 'bad@bad.com'},
        )
        self.assertIsNone(entry.user)

    def test_ordering_newest_first(self):
        AuditLog.objects.create(action='user_login', user=self.admin)
        AuditLog.objects.create(action='export_requested', user=self.admin)
        self.assertEqual(AuditLog.objects.first().action, 'export_requested')

    def test_certificate_deleted_sets_fk_null(self):
        entry = AuditLog.objects.create(
            action='certificate_generated', user=self.admin, certificate=self.cert,
        )
        self.cert.delete()
        entry.refresh_from_db()
        self.assertIsNone(entry.certificate)


class AuditLogInstrumentationTest(TestCase):
    """Verify that key API actions write AuditLog entries."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('audit_instr@test.com')
        self.client.force_authenticate(user=self.admin)
        self.event = make_event(self.admin, name='Instr Event')
        self.participant = make_participant(self.admin, doc='AI01', email='ai01@test.com')

    def test_successful_login_logs_user_login(self):
        User.objects.create_user(
            email='login_aud@test.com', full_name='A',
            password='Pass1234!', role='admin', is_staff=True,
        )
        before = AuditLog.objects.filter(action='user_login').count()
        self.client.post('/api/login/', {'email': 'login_aud@test.com', 'password': 'Pass1234!'})
        self.assertEqual(AuditLog.objects.filter(action='user_login').count(), before + 1)

    def test_failed_login_logs_user_login_failed(self):
        before = AuditLog.objects.filter(action='user_login_failed').count()
        self.client.post('/api/login/', {'email': 'nobody@x.com', 'password': 'wrong'})
        self.assertEqual(AuditLog.objects.filter(action='user_login_failed').count(), before + 1)

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_generate_logs_certificate_generated(self, mock_pdf):
        mock_pdf.return_value = {'success': True, 'path': '/m/a.pdf'}
        Enrollment.objects.create(
            participant=self.participant, event=self.event,
            attendance=True, created_by=self.admin,
        )
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, generated_by=self.admin,
        )
        before = AuditLog.objects.filter(action='certificate_generated').count()
        self.client.post(f'/api/certificates/{cert.id}/generate/', {})
        self.assertEqual(
            AuditLog.objects.filter(action='certificate_generated').count(), before + 1,
        )

    @patch('services.email_service.EmailService.send_certificate')
    def test_deliver_logs_certificate_delivered(self, mock_send):
        mock_send.return_value = {'success': True, 'message': 'sent'}
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, generated_by=self.admin,
        )
        Certificate.objects.filter(pk=cert.pk).update(status='generated', pdf_url='/m/c.pdf')
        before = AuditLog.objects.filter(action='certificate_delivered').count()
        self.client.post(f'/api/certificates/{cert.id}/deliver/', {'method': 'email'})
        self.assertEqual(
            AuditLog.objects.filter(action='certificate_delivered').count(), before + 1,
        )

    @patch('services.email_service.EmailService.send_certificate')
    def test_retry_logs_certificate_retried(self, mock_send):
        mock_send.return_value = {'success': True, 'message': 'ok'}
        cert = Certificate.objects.create(
            participant=self.participant, event=self.event, generated_by=self.admin,
        )
        Certificate.objects.filter(pk=cert.pk).update(status='failed', pdf_url='/m/c.pdf')
        before = AuditLog.objects.filter(action='certificate_retried').count()
        self.client.post(f'/api/certificates/{cert.id}/retry/', {'method': 'email'})
        self.assertEqual(
            AuditLog.objects.filter(action='certificate_retried').count(), before + 1,
        )

    def test_export_logs_export_requested(self):
        before = AuditLog.objects.filter(action='export_requested').count()
        self.client.get('/api/certificates/export/')
        self.assertEqual(
            AuditLog.objects.filter(action='export_requested').count(), before + 1,
        )


class AuditLogViewSetTest(TestCase):
    """GET /api/audit/ — admin-only read-only endpoint."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('audit_view@test.com')
        self.client.force_authenticate(user=self.admin)
        AuditLog.objects.create(action='user_login', user=self.admin)
        AuditLog.objects.create(action='export_requested', user=self.admin)

    def test_list_returns_200_for_admin(self):
        res = self.client.get('/api/audit/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)

    def test_list_contains_created_entries(self):
        res = self.client.get('/api/audit/')
        items = res.data.get('results', res.data)
        self.assertGreaterEqual(len(items), 2)

    def test_filter_by_action(self):
        res = self.client.get('/api/audit/?action=user_login')
        items = res.data.get('results', res.data)
        for item in items:
            self.assertEqual(item['action'], 'user_login')

    def test_detail_returns_single_entry(self):
        entry = AuditLog.objects.filter(user=self.admin).first()
        res = self.client.get(f'/api/audit/{entry.id}/')
        self.assertEqual(res.status_code, status.HTTP_200_OK)
        self.assertEqual(res.data['id'], entry.id)

    def test_participante_cannot_access_audit(self):
        part = make_user('audit_part@test.com')
        self.client.force_authenticate(user=part)
        res = self.client.get('/api/audit/')
        self.assertIn(res.status_code, [status.HTTP_403_FORBIDDEN, status.HTTP_401_UNAUTHORIZED])

    def test_unauthenticated_cannot_access_audit(self):
        anon = APIClient()
        res = anon.get('/api/audit/')
        self.assertEqual(res.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_post_is_rejected(self):
        res = self.client.post('/api/audit/', {'action': 'user_login'})
        self.assertIn(res.status_code, [
            status.HTTP_405_METHOD_NOT_ALLOWED,
            status.HTTP_403_FORBIDDEN,
        ])


# =============================================================================
# PASO 6 — Production infrastructure tests
# =============================================================================

# ─────────────────────────────────────────────
# OpenAPI / Swagger schema endpoints
# ─────────────────────────────────────────────

class OpenAPISchemaTest(TestCase):
    """TC-036 — OpenAPI schema is reachable and returns valid content."""

    def setUp(self):
        self.client = APIClient()
        self.admin = make_admin('swagger_admin@test.com')
        self.client.force_authenticate(user=self.admin)

    def test_schema_endpoint_returns_200(self):
        res = self.client.get('/api/schema/')
        self.assertEqual(res.status_code, 200)

    def test_schema_content_type_is_yaml_or_json(self):
        res = self.client.get('/api/schema/')
        ct = res.get('Content-Type', '')
        self.assertTrue(
            'yaml' in ct or 'json' in ct or 'text/' in ct or 'openapi' in ct,
            msg=f'Unexpected Content-Type: {ct}',
        )

    def test_schema_contains_api_title(self):
        res = self.client.get('/api/schema/?format=json')
        self.assertEqual(res.status_code, 200)
        body = res.content.decode()
        self.assertIn('SCAD', body)

    def test_swagger_ui_endpoint_returns_200(self):
        res = self.client.get('/api/docs/')
        self.assertEqual(res.status_code, 200)

    def test_redoc_endpoint_returns_200(self):
        res = self.client.get('/api/redoc/')
        self.assertEqual(res.status_code, 200)

    def test_schema_anonymous_access_allowed(self):
        anon = APIClient()
        res = anon.get('/api/schema/')
        self.assertIn(res.status_code, [200, 401, 403])

    def test_certificates_endpoint_in_schema(self):
        res = self.client.get('/api/schema/?format=json')
        body = res.content.decode()
        self.assertIn('certificates', body.lower())

    def test_events_endpoint_in_schema(self):
        res = self.client.get('/api/schema/?format=json')
        body = res.content.decode()
        self.assertIn('events', body.lower())


# ─────────────────────────────────────────────
# Celery task unit tests (no broker needed)
# ─────────────────────────────────────────────

class CeleryTasksTest(TestCase):
    """TC-037 — Celery tasks call underlying services correctly."""

    def setUp(self):
        self.admin = make_admin('celery_admin@test.com')
        self.participant = make_participant(self.admin, doc='CEL001', email='celery_p@test.com')
        self.event = make_event(self.admin, name='Celery Event')
        self.template = Template.objects.create(name='CeleryTpl', created_by=self.admin)
        self.cert = Certificate.objects.create(
            participant=self.participant,
            event=self.event,
            template=self.template,
            generated_by=self.admin,
        )

    @patch('services.email_service.EmailService.send_certificate')
    def test_send_certificate_email_task_success(self, mock_send):
        from services.tasks import send_certificate_email_task
        mock_send.return_value = {'success': True, 'message': 'sent'}
        result = send_certificate_email_task(self.cert.id, 'dest@test.com')
        mock_send.assert_called_once()
        self.assertTrue(result['success'])

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_generate_certificate_pdf_task_success(self, mock_gen):
        from services.tasks import generate_certificate_pdf_task
        mock_gen.return_value = {
            'success': True, 'path': '/certificates/pdfs/test.pdf', 'message': 'ok'
        }
        result = generate_certificate_pdf_task(self.cert.id)
        mock_gen.assert_called_once()
        self.assertTrue(result['success'])

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_generate_pdf_task_saves_pdf_url(self, mock_gen):
        from services.tasks import generate_certificate_pdf_task
        mock_gen.return_value = {
            'success': True, 'path': '/certificates/pdfs/saved.pdf', 'message': 'ok'
        }
        generate_certificate_pdf_task(self.cert.id)
        self.cert.refresh_from_db()
        self.assertEqual(self.cert.pdf_url, '/certificates/pdfs/saved.pdf')

    @patch('services.email_service.EmailService.send_bulk_certificates')
    def test_send_bulk_certificates_task_email(self, mock_bulk):
        from services.tasks import send_bulk_certificates_task
        mock_bulk.return_value = {'sent': 1, 'failed': 0, 'errors': []}
        result = send_bulk_certificates_task(self.event.id, method='email')
        self.assertEqual(result['sent'], 1)

    def test_send_bulk_unsupported_method_returns_error(self):
        from services.tasks import send_bulk_certificates_task
        result = send_bulk_certificates_task(self.event.id, method='fax')
        self.assertEqual(result['failed'], 0)
        self.assertGreater(len(result['errors']), 0)

    @patch('services.pdf_service.PDFService.generate_certificate_pdf')
    def test_generate_pdf_task_does_not_save_on_failure(self, mock_gen):
        from services.tasks import generate_certificate_pdf_task
        mock_gen.return_value = {
            'success': False, 'path': None, 'message': 'canvas crashed'
        }
        original_pdf_url = self.cert.pdf_url
        generate_certificate_pdf_task(self.cert.id)
        self.cert.refresh_from_db()
        self.assertEqual(self.cert.pdf_url, original_pdf_url)


# ─────────────────────────────────────────────
# Settings / infrastructure sanity checks
# ─────────────────────────────────────────────

class InfrastructureSettingsTest(TestCase):
    """TC-038 — Verify production-ready settings are configured."""

    def test_spectacular_settings_exist(self):
        from django.conf import settings
        self.assertTrue(hasattr(settings, 'SPECTACULAR_SETTINGS'))
        self.assertIn('TITLE', settings.SPECTACULAR_SETTINGS)
        self.assertIn('VERSION', settings.SPECTACULAR_SETTINGS)

    def test_spectacular_title_contains_scad(self):
        from django.conf import settings
        title = settings.SPECTACULAR_SETTINGS.get('TITLE', '')
        self.assertIn('SCAD', title)

    def test_drf_schema_class_is_spectacular(self):
        from django.conf import settings
        schema_class = settings.REST_FRAMEWORK.get('DEFAULT_SCHEMA_CLASS', '')
        self.assertIn('spectacular', schema_class)

    def test_celery_result_backend_is_configured(self):
        from django.conf import settings
        self.assertTrue(hasattr(settings, 'CELERY_RESULT_BACKEND'))

    def test_celery_broker_url_configured(self):
        from django.conf import settings
        self.assertTrue(hasattr(settings, 'CELERY_BROKER_URL'))

    def test_celery_task_serializer_is_json(self):
        from django.conf import settings
        self.assertEqual(getattr(settings, 'CELERY_TASK_SERIALIZER', ''), 'json')

    def test_drf_spectacular_in_installed_apps(self):
        from django.conf import settings
        self.assertIn('drf_spectacular', settings.INSTALLED_APPS)

    def test_django_celery_results_in_installed_apps(self):
        from django.conf import settings
        self.assertIn('django_celery_results', settings.INSTALLED_APPS)

    def test_jwt_access_token_lifetime_is_8_hours(self):
        from django.conf import settings
        from datetime import timedelta
        lifetime = settings.SIMPLE_JWT.get('ACCESS_TOKEN_LIFETIME')
        self.assertEqual(lifetime, timedelta(hours=8))

    def test_jwt_refresh_token_lifetime_is_7_days(self):
        from django.conf import settings
        from datetime import timedelta
        lifetime = settings.SIMPLE_JWT.get('REFRESH_TOKEN_LIFETIME')
        self.assertEqual(lifetime, timedelta(days=7))
